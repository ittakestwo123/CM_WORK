from __future__ import annotations

from calendar import monthrange
from datetime import datetime, timedelta
from io import BytesIO
import re
import time
from typing import Any
from uuid import uuid4

try:
    import psutil
except Exception:  # pragma: no cover - fallback when psutil is unavailable
    psutil = None
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from fastapi import Depends, FastAPI, HTTPException, Request, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from sqlalchemy import and_, case, func, inspect, select, or_, update
from sqlalchemy.exc import OperationalError
from sqlalchemy.orm import Session, joinedload

from .auth import create_access_token, hash_password, verify_password
from .database import Base, engine, get_db
from .dependencies import get_current_user, require_roles
from .models import (
    AuditLog,
    CustomRole,
    EmploymentReport,
    Enterprise,
    FilingStatus,
    Notice,
    NoticeRead,
    NoticeScope,
    NationalPushApiConfig,
    NationalPushLog,
    NationalPushRecord,
    OperationLog,
    ReportStatus,
    ReportVersion,
    SystemMonitorSnapshot,
    SurveyPeriod,
    SurveyPeriodStatus,
    SurveyPeriodType,
    User,
    UserRole,
)
from .schemas import (
    AuditLogOut,
    CompareCityOut,
    CompareMetricOut,
    CitySummaryItemOut,
    FilingOut,
    FilingPayload,
    FilingReviewIn,
    LoginIn,
    NoticeCreateIn,
    NationalPushApiConfigIn,
    NationalPushApiConfigOut,
    NationalPushExecuteIn,
    NationalPushLogOut,
    NationalPushPreviewOut,
    NationalPushRecordOut,
    NationalPushResultOut,
    NoticeOut,
    NoticeUpdateIn,
    PasswordChangeIn,
    ProvinceCompareOut,
    ProvinceMultiDimOut,
    ProvinceSamplingOut,
    ProvinceSummaryOut,
    ProvinceTrendOut,
    ReportExportRowOut,
    ReportOut,
    ReportModifyIn,
    ReportPayload,
    ReportReviewIn,
    ReportVersionOut,
    SamplingCityOut,
    MultiDimItemOut,
    SystemMonitorOut,
    SystemMonitorSnapshotOut,
    SystemRoleCreateIn,
    SystemRoleDeleteOut,
    SystemRoleOut,
    SystemRoleUpdateIn,
    SystemUserCreateIn,
    SystemUserOut,
    SystemUserRoleChangeIn,
    SystemUserUpdateIn,
    OperationLogOut,
    SurveyPeriodOut,
    SurveyPeriodUpsertIn,
    TokenOut,
    TrendPointOut,
    UserOut,
)


app = FastAPI(title="Employment Survey API", version="0.2.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

DASHBOARD_BY_ROLE: dict[UserRole, str] = {
    UserRole.ENTERPRISE: "/dashboard/enterprise",
    UserRole.CITY: "/dashboard/city",
    UserRole.PROVINCE: "/dashboard/province",
}

DECREASE_REASON_MAP: dict[str, set[str]] = {
    "生产经营调整": {"订单减少", "生产线调整", "其他"},
    "组织结构优化": {"组织精简", "岗位合并", "其他"},
    "员工主动流动": {"个人原因离职", "退休", "其他"},
    "其他": {"其他"},
}

EDITABLE_REPORT_STATUSES = {
    ReportStatus.DRAFT,
    ReportStatus.CITY_REJECTED,
    ReportStatus.PROVINCE_REJECTED,
}

PROVINCE_ANALYTIC_STATUSES = {
    ReportStatus.CITY_APPROVED,
    ReportStatus.PROVINCE_REJECTED,
    ReportStatus.PROVINCE_APPROVED,
    ReportStatus.PROVINCE_SUBMITTED,
}

ROLE_PERMISSIONS: dict[UserRole, list[str]] = {
    UserRole.ENTERPRISE: ["企业备案", "企业填报", "企业查询", "通知浏览"],
    UserRole.CITY: ["市级审核", "市级通知发布"],
    UserRole.PROVINCE: ["省级审核", "省级通知发布", "系统管理", "导出与对外推送"],
}

ROLE_NAME_MAP: dict[UserRole, str] = {
    UserRole.ENTERPRISE: "企业系统角色",
    UserRole.CITY: "市级系统角色",
    UserRole.PROVINCE: "省级系统角色",
}

LOCK_MAX_FAILURES = 5
LOCK_MINUTES = 10
NATIONAL_PUSH_MAX_RETRIES = 3
MONITOR_SNAPSHOT_KEEP = 20


def _normalize_permissions(values: list[str]) -> list[str]:
    ordered: list[str] = []
    for raw in values:
        item = raw.strip()
        if not item:
            continue
        if item not in ordered:
            ordered.append(item)
    return ordered


def _system_user_to_out(item: User) -> SystemUserOut:
    return SystemUserOut(
        id=item.id,
        username=item.username,
        role=item.role,
        custom_role_id=item.custom_role_id,
        custom_role_name=item.custom_role.name if item.custom_role else None,
        is_active=item.is_active,
        is_activated=item.is_activated,
        login_fail_count=item.login_fail_count,
        locked_until=item.locked_until,
        city_code=item.city_code,
        enterprise_id=item.enterprise_id,
        created_at=item.created_at,
    )


def _resolve_custom_role(
    db: Session,
    custom_role_id: int | None,
    expected_scope: UserRole,
) -> CustomRole | None:
    if custom_role_id is None:
        return None

    role_item = db.get(CustomRole, custom_role_id)
    if role_item is None or role_item.is_builtin:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="自定义角色不存在")
    if role_item.scope != expected_scope:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="自定义角色与用户基础角色不匹配")
    return role_item


def _build_system_roles(db: Session) -> list[SystemRoleOut]:
    output: list[SystemRoleOut] = []
    for role, permissions in ROLE_PERMISSIONS.items():
        assigned_count = db.scalar(select(func.count(User.id)).where(User.role == role)) or 0
        output.append(
            SystemRoleOut(
                id=f"builtin:{role.value}",
                name=ROLE_NAME_MAP.get(role, role.value),
                scope=role,
                is_builtin=True,
                assigned_user_count=assigned_count,
                permissions=permissions,
            )
        )

    custom_roles = list(db.scalars(select(CustomRole).order_by(CustomRole.created_at.desc())).all())
    for item in custom_roles:
        assigned_count = db.scalar(select(func.count(User.id)).where(User.custom_role_id == item.id)) or 0
        output.append(
            SystemRoleOut(
                id=f"custom:{item.id}",
                name=item.name,
                scope=item.scope,
                is_builtin=False,
                assigned_user_count=assigned_count,
                permissions=item.permissions or [],
            )
        )
    return output


def _custom_role_to_out(db: Session, item: CustomRole) -> SystemRoleOut:
    assigned_count = db.scalar(select(func.count(User.id)).where(User.custom_role_id == item.id)) or 0
    return SystemRoleOut(
        id=f"custom:{item.id}",
        name=item.name,
        scope=item.scope,
        is_builtin=False,
        assigned_user_count=assigned_count,
        permissions=item.permissions or [],
    )


def _ensure_schema_compatible() -> None:
    inspector = inspect(engine)
    try:
        report_columns = {col["name"] for col in inspector.get_columns("employment_reports")}
        user_columns = {col["name"] for col in inspector.get_columns("users")}
    except Exception:
        return

    report_required = {"is_deleted", "deleted_at"}
    user_required = {"is_activated", "login_fail_count", "locked_until", "custom_role_id"}
    role_table_ready = inspector.has_table("custom_roles")
    if not report_required.issubset(report_columns) or not user_required.issubset(user_columns) or not role_table_ready:
        Base.metadata.drop_all(bind=engine)
        Base.metadata.create_all(bind=engine)


def _report_is_active_filter():
    return EmploymentReport.is_deleted.is_(False)


@app.on_event("startup")
def startup() -> None:
    Base.metadata.create_all(bind=engine)
    _ensure_schema_compatible()
    with Session(engine) as db:
        try:
            _seed_data(db)
        except OperationalError:
            # Week2 introduces schema changes; recreate local DB when old schema exists.
            db.rollback()
            Base.metadata.drop_all(bind=engine)
            Base.metadata.create_all(bind=engine)
            _seed_data(db)


def _seed_data(db: Session) -> None:
    province_user = db.scalar(select(User).where(User.username == "province_admin"))
    if province_user is None:
        db.add(
            User(
                username="province_admin",
                password_hash=hash_password("Passw0rd!"),
                role=UserRole.PROVINCE,
                is_active=True,
                is_activated=True,
                city_code=None,
            )
        )
    else:
        province_user.password_hash = hash_password("Passw0rd!")
        province_user.role = UserRole.PROVINCE
        province_user.is_active = True
        province_user.is_activated = True
        province_user.city_code = None

    city_user = db.scalar(select(User).where(User.username == "city_reviewer"))
    if city_user is None:
        db.add(
            User(
                username="city_reviewer",
                password_hash=hash_password("Passw0rd!"),
                role=UserRole.CITY,
                is_active=True,
                is_activated=True,
                city_code="530100",
            )
        )
    else:
        city_user.password_hash = hash_password("Passw0rd!")
        city_user.role = UserRole.CITY
        city_user.is_active = True
        city_user.is_activated = True
        city_user.city_code = "530100"

    enterprise_user = db.scalar(select(User).where(User.username == "enterprise_user"))
    if enterprise_user is None:
        enterprise = Enterprise(
            name="示例企业",
            city_code="530100",
            city_name="昆明市",
            filing_status=FilingStatus.NOT_SUBMITTED,
        )
        db.add(enterprise)
        db.flush()

        db.add(
            User(
                username="enterprise_user",
                password_hash=hash_password("Passw0rd!"),
                role=UserRole.ENTERPRISE,
                is_active=True,
                is_activated=True,
                enterprise_id=enterprise.id,
                city_code=enterprise.city_code,
            )
        )
    else:
        enterprise_user.password_hash = hash_password("Passw0rd!")
        enterprise_user.role = UserRole.ENTERPRISE
        enterprise_user.is_active = True
        enterprise_user.is_activated = True
        if enterprise_user.enterprise_id is None:
            enterprise = Enterprise(
                name="示例企业",
                city_code=enterprise_user.city_code or "530100",
                city_name="昆明市",
                filing_status=FilingStatus.NOT_SUBMITTED,
            )
            db.add(enterprise)
            db.flush()
            enterprise_user.enterprise_id = enterprise.id
        else:
            enterprise = db.get(Enterprise, enterprise_user.enterprise_id)
            if enterprise is not None:
                if not enterprise.city_code:
                    enterprise.city_code = enterprise_user.city_code or "530100"
                if not enterprise.city_name:
                    enterprise.city_name = "昆明市"

    disabled_user = db.scalar(select(User).where(User.username == "disabled_user"))
    if disabled_user is None:
        db.add(
            User(
                username="disabled_user",
                password_hash=hash_password("Passw0rd!"),
                role=UserRole.ENTERPRISE,
                is_active=False,
                is_activated=True,
                city_code="530100",
            )
        )
    else:
        disabled_user.password_hash = hash_password("Passw0rd!")
        disabled_user.role = UserRole.ENTERPRISE
        disabled_user.is_active = False
        disabled_user.is_activated = True
        disabled_user.city_code = "530100"

    inactive_user = db.scalar(select(User).where(User.username == "inactive_user"))
    if inactive_user is None:
        db.add(
            User(
                username="inactive_user",
                password_hash=hash_password("Passw0rd!"),
                role=UserRole.ENTERPRISE,
                is_active=True,
                is_activated=False,
                city_code="530100",
            )
        )
    else:
        inactive_user.password_hash = hash_password("Passw0rd!")
        inactive_user.role = UserRole.ENTERPRISE
        inactive_user.is_active = True
        inactive_user.is_activated = False
        inactive_user.city_code = "530100"

    db.flush()
    _seed_survey_periods(db)
    db.commit()


def _seed_survey_periods(db: Session) -> None:
    if db.scalar(select(SurveyPeriod.id).limit(1)) is not None:
        return

    year = datetime.utcnow().year
    defaults: list[SurveyPeriod] = []

    for month_no in (1, 2, 3):
        defaults.append(
            SurveyPeriod(
                period_code=f"{year}-{month_no:02d}-H1",
                period_name=f"{year}年{month_no}月上半月",
                start_time=datetime(year, month_no, 1, 0, 0, 0),
                end_time=datetime(year, month_no, 15, 23, 59, 59),
                period_type=SurveyPeriodType.HALF_MONTH,
                month_no=month_no,
                half_no=1,
                status=SurveyPeriodStatus.ENABLED,
            )
        )
        end_day = monthrange(year, month_no)[1]
        defaults.append(
            SurveyPeriod(
                period_code=f"{year}-{month_no:02d}-H2",
                period_name=f"{year}年{month_no}月下半月",
                start_time=datetime(year, month_no, 16, 0, 0, 0),
                end_time=datetime(year, month_no, end_day, 23, 59, 59),
                period_type=SurveyPeriodType.HALF_MONTH,
                month_no=month_no,
                half_no=2,
                status=SurveyPeriodStatus.ENABLED,
            )
        )

    current_month = datetime.utcnow().month
    if current_month in {1, 2, 3}:
        current_month = 4
    end_day = monthrange(year, current_month)[1]
    defaults.append(
        SurveyPeriod(
            period_code=f"{year}-{current_month:02d}-M1",
            period_name=f"{year}年{current_month}月月报",
            start_time=datetime(year, current_month, 1, 0, 0, 0),
            end_time=datetime(year, current_month, end_day, 23, 59, 59),
            period_type=SurveyPeriodType.MONTH,
            month_no=current_month,
            half_no=None,
            status=SurveyPeriodStatus.ENABLED,
        )
    )

    for period in defaults:
        db.add(period)


def _enterprise_snapshot(item: Enterprise) -> dict[str, Any]:
    return {
        "id": item.id,
        "organization_code": item.organization_code,
        "name": item.name,
        "nature": item.nature,
        "industry": item.industry,
        "business_scope": item.business_scope,
        "contact_person": item.contact_person,
        "contact_address": item.contact_address,
        "postal_code": item.postal_code,
        "phone": item.phone,
        "fax": item.fax,
        "email": item.email,
        "city_code": item.city_code,
        "city_name": item.city_name,
        "filing_status": item.filing_status.value,
        "filing_reject_reason": item.filing_reject_reason,
    }


def _report_snapshot(item: EmploymentReport) -> dict[str, Any]:
    return {
        "id": item.id,
        "enterprise_id": item.enterprise_id,
        "survey_period_id": item.survey_period_id,
        "period_code": item.period_code,
        "base_employment": item.base_employment,
        "survey_employment": item.survey_employment,
        "decrease_type": item.decrease_type,
        "decrease_reason": item.decrease_reason,
        "decrease_reason_detail": item.decrease_reason_detail,
        "other_note": item.other_note,
        "decrease_count": item.decrease_count,
        "status": item.status.value,
        "last_submitted_at": item.last_submitted_at.isoformat() if item.last_submitted_at else None,
    }


def _report_to_out(item: EmploymentReport) -> ReportOut:
    period = item.survey_period
    enterprise = item.enterprise
    return ReportOut(
        id=item.id,
        enterprise_id=item.enterprise_id,
        enterprise_name=enterprise.name if enterprise else None,
        region=enterprise.city_name if enterprise else None,
        survey_period_id=item.survey_period_id,
        period_code=item.period_code,
        period_name=period.period_name if period else None,
        period_type=period.period_type if period else None,
        base_employment=item.base_employment,
        survey_employment=item.survey_employment,
        decrease_type=item.decrease_type,
        decrease_reason=item.decrease_reason,
        decrease_reason_detail=item.decrease_reason_detail,
        other_note=item.other_note,
        decrease_count=item.decrease_count,
        status=item.status,
        last_submitted_at=item.last_submitted_at,
        city_reviewed_at=item.city_reviewed_at,
        province_reviewed_at=item.province_reviewed_at,
    )


def _write_audit_log(
    db: Session,
    request: Request,
    operator: User,
    action_type: str,
    target_type: str,
    target_id: str,
    reason: str | None,
    before_state: dict[str, Any] | None,
    after_state: dict[str, Any] | None,
) -> None:
    ip_address = request.client.host if request.client else None
    log = AuditLog(
        operator_id=operator.id,
        operator_role=operator.role.value,
        action_type=action_type,
        target_type=target_type,
        target_id=target_id,
        reason=reason,
        before_state=before_state,
        after_state=after_state,
        ip_address=ip_address,
    )
    db.add(log)

    op_log = OperationLog(
        user_id=operator.id,
        user_name=operator.username,
        role=operator.role.value,
        operation_type=action_type,
        target_type=target_type,
        target_id=target_id,
        old_value=before_state,
        new_value=after_state,
        reason=reason,
        ip_address=ip_address,
    )
    db.add(op_log)


def _save_report_version(db: Session, report: EmploymentReport, operator: User, action_type: str) -> None:
    db.flush()
    current_no = db.scalar(
        select(func.max(ReportVersion.version_no)).where(ReportVersion.report_id == report.id)
    )
    next_no = (current_no or 0) + 1
    db.add(
        ReportVersion(
            report_id=report.id,
            version_no=next_no,
            action_type=action_type,
            operator_id=operator.id,
            snapshot=_report_snapshot(report),
        )
    )


def _get_or_create_enterprise(db: Session, user: User) -> Enterprise:
    enterprise = db.get(Enterprise, user.enterprise_id) if user.enterprise_id else None
    if enterprise is not None:
        return enterprise

    enterprise = Enterprise(
        created_by_user_id=user.id,
        filing_status=FilingStatus.NOT_SUBMITTED,
        city_code=user.city_code or "530100",
        city_name="昆明市",
    )
    db.add(enterprise)
    db.flush()

    user.enterprise_id = enterprise.id
    db.flush()
    return enterprise


def _get_period_by_code(db: Session, period_code: str) -> SurveyPeriod:
    period = db.scalar(select(SurveyPeriod).where(SurveyPeriod.period_code == period_code))
    if period is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="调查期不存在")
    return period


def _get_report_with_relations(db: Session, report_id: int) -> EmploymentReport:
    stmt = (
        select(EmploymentReport)
        .options(joinedload(EmploymentReport.survey_period), joinedload(EmploymentReport.enterprise))
        .where(EmploymentReport.id == report_id, _report_is_active_filter())
    )
    report = db.scalar(stmt)
    if report is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="报表不存在")
    return report


def _check_enterprise_can_report(enterprise: Enterprise, period: SurveyPeriod) -> None:
    if enterprise.filing_status != FilingStatus.APPROVED:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="企业未备案通过，不能填报")

    if period.status != SurveyPeriodStatus.ENABLED:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="调查期未启用，不能填报")

    now = datetime.utcnow()
    if not (period.start_time <= now <= period.end_time):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="当前不在所选调查期内，不能填报")


def _validate_submission(payload: ReportPayload) -> None:
    if payload.survey_employment >= payload.base_employment:
        return

    if not payload.decrease_type:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="就业人数减少时必须填写减少类型")
    if not payload.decrease_reason:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="就业人数减少时必须填写主要原因")
    if not payload.decrease_reason_detail:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="就业人数减少时必须填写主要原因说明")

    valid_reasons = DECREASE_REASON_MAP.get(payload.decrease_type)
    if valid_reasons is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="减少类型不在允许范围内")

    if payload.decrease_reason not in valid_reasons:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="减少类型与主要原因不匹配")

    if payload.decrease_type == "其他" or payload.decrease_reason == "其他":
        if not payload.other_note:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="选择其他时必须填写补充说明")


def _validate_period_payload(
    db: Session,
    payload: SurveyPeriodUpsertIn,
    exclude_period_id: int | None = None,
) -> None:
    half_pattern = re.compile(r"^\d{4}-(0[1-9]|1[0-2])-H[12]$")
    month_pattern = re.compile(r"^\d{4}-(0[1-9]|1[0-2])-M1$")

    if payload.start_time >= payload.end_time:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="调查期开始时间必须早于结束时间")

    if payload.start_time.year != payload.end_time.year or payload.start_time.month != payload.end_time.month:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="调查期必须位于同一年同一月份")

    if payload.start_time.month != payload.month_no:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="month_no 与开始时间月份不一致")

    if payload.month_no in {1, 2, 3}:
        if payload.period_type != SurveyPeriodType.HALF_MONTH:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="1/2/3月必须使用半月调查期")
        if payload.half_no not in {1, 2}:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="1/2/3月必须指定 half_no 为 1 或 2")
        if not half_pattern.fullmatch(payload.period_code):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="半月调查期编码格式应为 YYYY-MM-H1/H2")
    else:
        if payload.period_type != SurveyPeriodType.MONTH:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="4-12月仅允许创建月度调查期")
        if payload.half_no is not None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="月度调查期 half_no 必须为空")
        if not month_pattern.fullmatch(payload.period_code):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="月度调查期编码格式应为 YYYY-MM-M1")

    duplicate_code_stmt = select(SurveyPeriod).where(SurveyPeriod.period_code == payload.period_code)
    duplicate_code = db.scalar(duplicate_code_stmt)
    if duplicate_code is not None and duplicate_code.id != exclude_period_id:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="period_code 已存在")

    all_periods = list(db.scalars(select(SurveyPeriod)).all())
    year_no = payload.start_time.year
    same_month_periods = [
        p
        for p in all_periods
        if p.id != exclude_period_id and p.start_time.year == year_no and p.month_no == payload.month_no
    ]

    if payload.month_no in {1, 2, 3}:
        if len(same_month_periods) >= 2:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="1/2/3月最多允许两个调查期")
        if any(p.half_no == payload.half_no for p in same_month_periods):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="该半月调查期已存在")
    else:
        if same_month_periods:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="4-12月每月仅允许一个调查期")

    for period in same_month_periods:
        overlaps = payload.start_time <= period.end_time and payload.end_time >= period.start_time
        if overlaps:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="调查期时间区间与现有记录重叠")


def _upsert_report_for_enterprise(
    db: Session,
    enterprise: Enterprise,
    payload: ReportPayload,
) -> tuple[EmploymentReport, dict[str, Any] | None]:
    period = _get_period_by_code(db, payload.period_code)
    _check_enterprise_can_report(enterprise, period)

    stmt = select(EmploymentReport).where(
        EmploymentReport.enterprise_id == enterprise.id,
        EmploymentReport.survey_period_id == period.id,
    )
    report = db.scalar(stmt)

    before_state: dict[str, Any] | None = None
    if report is None:
        report = EmploymentReport(
            enterprise_id=enterprise.id,
            survey_period_id=period.id,
            period_code=period.period_code,
        )
        db.add(report)
    else:
        if report.is_deleted:
            report.is_deleted = False
            report.deleted_at = None
        if report.status not in EDITABLE_REPORT_STATUSES:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="当前报表状态不允许修改")
        before_state = _report_snapshot(report)

    report.period_code = period.period_code
    report.base_employment = payload.base_employment
    report.survey_employment = payload.survey_employment
    report.decrease_type = payload.decrease_type
    report.decrease_reason = payload.decrease_reason
    report.decrease_reason_detail = payload.decrease_reason_detail
    report.other_note = payload.other_note

    if payload.survey_employment >= payload.base_employment:
        report.decrease_type = None
        report.decrease_reason = None
        report.decrease_reason_detail = None
        report.other_note = None

    return report, before_state


def _assert_city_scope(current_user: User, report: EmploymentReport) -> None:
    if not current_user.city_code:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="当前市级账号未绑定辖区")

    if report.enterprise is None or report.enterprise.city_code != current_user.city_code:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="仅可操作本辖区企业报表")


def _list_city_reports(db: Session, current_user: User, status_filter: ReportStatus | None) -> list[ReportOut]:
    if not current_user.city_code:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="当前市级账号未绑定辖区")

    effective_status = status_filter or ReportStatus.PENDING_CITY_REVIEW
    stmt = (
        select(EmploymentReport)
        .join(EmploymentReport.enterprise)
        .options(joinedload(EmploymentReport.enterprise), joinedload(EmploymentReport.survey_period))
        .where(
            Enterprise.city_code == current_user.city_code,
            EmploymentReport.status == effective_status,
            _report_is_active_filter(),
        )
        .order_by(EmploymentReport.updated_at.desc())
    )
    reports = list(db.scalars(stmt).all())
    return [_report_to_out(item) for item in reports]


def _list_province_reports(db: Session, status_filter: ReportStatus | None) -> list[ReportOut]:
    stmt = select(EmploymentReport).options(
        joinedload(EmploymentReport.enterprise),
        joinedload(EmploymentReport.survey_period),
    )
    if status_filter is not None:
        stmt = stmt.where(EmploymentReport.status == status_filter, _report_is_active_filter())
    else:
        stmt = stmt.where(
            _report_is_active_filter(),
            EmploymentReport.status.in_(
                [
                    ReportStatus.CITY_APPROVED,
                    ReportStatus.PROVINCE_REJECTED,
                    ReportStatus.PROVINCE_APPROVED,
                    ReportStatus.PROVINCE_SUBMITTED,
                ]
            )
        )
    stmt = stmt.order_by(EmploymentReport.updated_at.desc())
    reports = list(db.scalars(stmt).all())
    return [_report_to_out(item) for item in reports]


def _check_delete_allowed(report: EmploymentReport) -> tuple[bool, str | None]:
    if report.status == ReportStatus.PROVINCE_SUBMITTED:
        return False, "已上报部委数据禁止删除"
    if report.status == ReportStatus.PROVINCE_APPROVED:
        return False, "已参与汇总的数据禁止删除"
    return True, None


def _check_user_delete_allowed(db: Session, current_user: User, target: User) -> tuple[bool, str | None]:
    if target.id == current_user.id:
        return False, "不能删除当前登录用户"

    if target.username in {"province_admin", "city_reviewer", "enterprise_user"}:
        return False, "系统内置演示账号禁止删除"

    if target.role == UserRole.PROVINCE:
        province_active = db.scalar(
            select(func.count(User.id)).where(
                User.role == UserRole.PROVINCE,
                User.is_active.is_(True),
                User.id != target.id,
            )
        ) or 0
        if province_active < 1:
            return False, "至少保留 1 名启用状态的省级用户"

    if target.role == UserRole.CITY and target.city_code:
        pending_city = db.scalar(
            select(func.count(EmploymentReport.id))
            .join(EmploymentReport.enterprise)
            .where(
                Enterprise.city_code == target.city_code,
                EmploymentReport.status == ReportStatus.PENDING_CITY_REVIEW,
                _report_is_active_filter(),
            )
        ) or 0
        if pending_city > 0:
            return False, "该市级账号辖区内存在待市审报表，禁止删除"

    if target.role == UserRole.ENTERPRISE and target.enterprise_id:
        report_count = db.scalar(
            select(func.count(EmploymentReport.id)).where(
                EmploymentReport.enterprise_id == target.enterprise_id,
                _report_is_active_filter(),
            )
        ) or 0
        if report_count > 0:
            return False, "该企业存在历史报表数据，禁止删除"

    return True, None


def _notice_visible_filter(current_user: User):
    if current_user.role == UserRole.PROVINCE:
        return Notice.is_deleted.is_(False)

    if current_user.role == UserRole.CITY:
        return and_(
            Notice.is_deleted.is_(False),
            or_(
                Notice.scope == NoticeScope.PROVINCE,
                and_(Notice.scope == NoticeScope.CITY, Notice.city_code == current_user.city_code),
            ),
        )

    return and_(
        Notice.is_deleted.is_(False),
        or_(
            Notice.scope == NoticeScope.PROVINCE,
            and_(Notice.scope == NoticeScope.CITY, Notice.city_code == current_user.city_code),
        ),
    )


def _notice_to_out(notice: Notice, is_read: bool) -> NoticeOut:
    return NoticeOut(
        id=notice.id,
        title=notice.title,
        content=notice.content,
        scope=notice.scope,
        city_code=notice.city_code,
        publisher_name=notice.publisher_name,
        publisher_role=notice.publisher_role,
        read_count=notice.read_count,
        read=is_read,
        created_at=notice.created_at,
        updated_at=notice.updated_at,
    )


def _resolve_analytic_period(db: Session, period_code: str | None) -> SurveyPeriod:
    if period_code:
        return _get_period_by_code(db, period_code)

    stmt = (
        select(SurveyPeriod)
        .join(EmploymentReport, EmploymentReport.survey_period_id == SurveyPeriod.id)
        .where(EmploymentReport.status.in_(PROVINCE_ANALYTIC_STATUSES), _report_is_active_filter())
        .order_by(SurveyPeriod.start_time.desc())
        .limit(1)
    )
    period = db.scalar(stmt)
    if period is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="暂无可用于统计分析的数据")
    return period


def _province_city_summary_rows(db: Session, period_id: int) -> list[CitySummaryItemOut]:
    rows = _province_dimension_summary_rows(db, period_id, "region")
    return [
        CitySummaryItemOut(
            city=item["value"],
            enterprise_count=item["enterprise_count"],
            base_total=item["base_total"],
            survey_total=item["survey_total"],
            change_total=item["change_total"],
        )
        for item in rows
    ]


def _province_dimension_expr(dimension: str):
    if dimension == "region":
        return func.coalesce(Enterprise.city_name, Enterprise.city_code, "未知地区")
    if dimension == "nature":
        return func.coalesce(Enterprise.nature, "未填写")
    if dimension == "industry":
        return func.coalesce(Enterprise.industry, "未填写")
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="分析维度仅支持 region/nature/industry")


def _province_dimension_summary_rows(
    db: Session,
    period_id: int,
    dimension: str,
    city: str | None = None,
    nature: str | None = None,
    industry: str | None = None,
) -> list[dict[str, Any]]:
    dim_expr = _province_dimension_expr(dimension)
    decrease_expr = case(
        (
            EmploymentReport.base_employment > EmploymentReport.survey_employment,
            EmploymentReport.base_employment - EmploymentReport.survey_employment,
        ),
        else_=0,
    )

    stmt = (
        select(
            dim_expr.label("dimension_value"),
            func.count(func.distinct(EmploymentReport.enterprise_id)).label("enterprise_count"),
            func.coalesce(func.sum(EmploymentReport.base_employment), 0).label("base_total"),
            func.coalesce(func.sum(EmploymentReport.survey_employment), 0).label("survey_total"),
            func.coalesce(func.sum(decrease_expr), 0).label("decrease_total"),
        )
        .join(EmploymentReport.enterprise)
        .where(
            EmploymentReport.survey_period_id == period_id,
            EmploymentReport.status.in_(PROVINCE_ANALYTIC_STATUSES),
            _report_is_active_filter(),
        )
    )

    if city:
        city_q = city.strip()
        stmt = stmt.where(or_(Enterprise.city_name.contains(city_q), Enterprise.city_code.contains(city_q)))
    if nature:
        stmt = stmt.where(Enterprise.nature.contains(nature.strip()))
    if industry:
        stmt = stmt.where(Enterprise.industry.contains(industry.strip()))

    stmt = stmt.group_by(dim_expr).order_by(dim_expr.asc())
    rows = db.execute(stmt).all()

    output: list[dict[str, Any]] = []
    for row in rows:
        base_total = int(row.base_total or 0)
        survey_total = int(row.survey_total or 0)
        change_total = survey_total - base_total
        change_ratio = round((change_total / base_total) * 100, 2) if base_total else 0.0
        output.append(
            {
                "value": str(row.dimension_value),
                "enterprise_count": int(row.enterprise_count or 0),
                "base_total": base_total,
                "survey_total": survey_total,
                "change_total": change_total,
                "decrease_total": int(row.decrease_total or 0),
                "change_ratio": change_ratio,
            }
        )
    return output


def _ratio_text(value_a: int, value_b: int) -> str:
    if value_a == 0:
        return "-"
    return f"{((value_b - value_a) / value_a) * 100:+.1f}%"


def _mask_api_key(raw: str) -> str:
    clean = raw.strip()
    if len(clean) <= 8:
        return "*" * len(clean)
    return f"{clean[:4]}{'*' * (len(clean) - 8)}{clean[-4:]}"


def _display_filter_map(filters: dict[str, Any]) -> str:
    visible: list[str] = []
    for key, value in filters.items():
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        if isinstance(value, datetime):
            visible.append(f"{key}={value.strftime('%Y-%m-%d %H:%M:%S')}")
            continue
        if isinstance(value, ReportStatus):
            visible.append(f"{key}={value.value}")
            continue
        visible.append(f"{key}={value}")
    return "；".join(visible) if visible else "无"


def _build_xlsx_response(
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    query_filters: dict[str, Any],
    operator: User,
    file_name: str,
) -> Response:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导出结果"

    sheet.append([title])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
    sheet["A1"].font = Font(size=14, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")

    now_text = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append(["导出时间", now_text])
    sheet.append(["操作用户", operator.username])
    sheet.append(["查询条件", _display_filter_map(query_filters)])
    sheet.append([])

    sheet.append(headers)
    head_row = sheet.max_row
    fill = PatternFill("solid", fgColor="DCE6F1")
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=head_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        rendered = []
        for value in row:
            if isinstance(value, datetime):
                rendered.append(value.strftime("%Y-%m-%d %H:%M:%S"))
            elif isinstance(value, ReportStatus):
                rendered.append(value.value)
            else:
                rendered.append(value)
        sheet.append(rendered)

    for col in range(1, len(headers) + 1):
        max_len = 10
        for row in range(1, sheet.max_row + 1):
            value = sheet.cell(row=row, column=col).value
            text = "" if value is None else str(value)
            max_len = max(max_len, min(40, len(text) + 2))
        sheet.column_dimensions[sheet.cell(row=head_row, column=col).column_letter].width = max_len

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    content = stream.getvalue()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={file_name}"},
    )


def _clear_login_lock(user: User) -> None:
    user.login_fail_count = 0
    user.locked_until = None


def _mark_login_failure(user: User) -> bool:
    user.login_fail_count = (user.login_fail_count or 0) + 1
    if user.login_fail_count >= LOCK_MAX_FAILURES:
        user.locked_until = datetime.utcnow() + timedelta(minutes=LOCK_MINUTES)
        user.login_fail_count = 0
        return True
    return False


def _calc_warning_level(cpu_usage: int, memory_usage: int, disk_usage: int, pending_review_count: int, api_latency_ms: int) -> str:
    if cpu_usage >= 90 or memory_usage >= 90 or disk_usage >= 92 or api_latency_ms >= 1800:
        return "严重"
    if cpu_usage >= 75 or memory_usage >= 75 or disk_usage >= 80 or pending_review_count >= 50 or api_latency_ms >= 900:
        return "预警"
    return "正常"


def _append_push_log(
    db: Session,
    record: NationalPushRecord,
    attempt_no: int,
    status: str,
    status_code: int | None,
    message: str,
    detail: dict[str, Any] | None = None,
) -> None:
    db.add(
        NationalPushLog(
            record_id=record.id,
            batch_id=record.batch_id,
            attempt_no=attempt_no,
            status=status,
            status_code=status_code,
            message=message,
            detail=detail,
        )
    )


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.post("/auth/login", response_model=TokenOut)
def login(payload: LoginIn, db: Session = Depends(get_db)) -> TokenOut:
    user = db.scalar(select(User).where(User.username == payload.username))
    if user is None:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="用户名或密码错误")

    now = datetime.utcnow()
    if user.locked_until is not None and user.locked_until > now:
        raise HTTPException(
            status_code=status.HTTP_423_LOCKED,
            detail=f"连续登录失败过多，账号已锁定至 {user.locked_until.strftime('%Y-%m-%d %H:%M:%S')}",
        )
    if not user.is_activated:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="账号未激活")
    if not user.is_active:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="账号已停用")

    if not verify_password(payload.password, user.password_hash):
        locked = _mark_login_failure(user)
        db.commit()
        if locked:
            raise HTTPException(
                status_code=status.HTTP_423_LOCKED,
                detail=f"连续 {LOCK_MAX_FAILURES} 次登录失败，账号已锁定 {LOCK_MINUTES} 分钟",
            )
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="用户名或密码错误")

    _clear_login_lock(user)
    db.commit()

    access_token = create_access_token(user)
    return TokenOut(
        access_token=access_token,
        role=user.role,
        dashboard=DASHBOARD_BY_ROLE[user.role],
    )


@app.get("/auth/me", response_model=UserOut)
def me(current_user: User = Depends(get_current_user)) -> UserOut:
    if current_user.role == UserRole.ENTERPRISE:
        region = current_user.enterprise.city_name if current_user.enterprise else (current_user.city_code or "未知地区")
        name = current_user.enterprise.name if current_user.enterprise and current_user.enterprise.name else current_user.username
    elif current_user.role == UserRole.CITY:
        region = current_user.city_code or "未配置辖区"
        name = f"市级审核员-{current_user.username}"
    else:
        region = "云南省"
        name = f"省级管理员-{current_user.username}"

    return UserOut(
        id=current_user.id,
        username=current_user.username,
        name=name,
        region=region,
        role=current_user.role,
        is_active=current_user.is_active,
        is_activated=current_user.is_activated,
        login_fail_count=current_user.login_fail_count,
        locked_until=current_user.locked_until,
        city_code=current_user.city_code,
        enterprise_id=current_user.enterprise_id,
    )


@app.post("/auth/change-password")
def change_password(
    payload: PasswordChangeIn,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    if not verify_password(payload.old_password, current_user.password_hash):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="原密码错误")
    if payload.old_password == payload.new_password:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="新密码不能与原密码相同")

    current_user.password_hash = hash_password(payload.new_password)
    _clear_login_lock(current_user)
    db.flush()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="auth_change_password",
        target_type="user",
        target_id=str(current_user.id),
        reason=None,
        before_state=None,
        after_state={"password_changed": True},
    )
    db.commit()
    return {"changed": True}


@app.get("/survey-periods", response_model=list[SurveyPeriodOut])
def list_survey_periods(
    month_no: int | None = None,
    period_type: SurveyPeriodType | None = None,
    status_value: SurveyPeriodStatus | None = None,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[SurveyPeriodOut]:
    stmt = select(SurveyPeriod).order_by(SurveyPeriod.start_time.desc())
    if month_no is not None:
        stmt = stmt.where(SurveyPeriod.month_no == month_no)
    if period_type is not None:
        stmt = stmt.where(SurveyPeriod.period_type == period_type)
    if status_value is not None:
        stmt = stmt.where(SurveyPeriod.status == status_value)
    return list(db.scalars(stmt).all())


@app.get("/survey-periods/current", response_model=list[SurveyPeriodOut])
def list_current_periods(
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[SurveyPeriodOut]:
    now = datetime.utcnow()
    stmt = (
        select(SurveyPeriod)
        .where(
            SurveyPeriod.start_time <= now,
            SurveyPeriod.end_time >= now,
            SurveyPeriod.status == SurveyPeriodStatus.ENABLED,
        )
        .order_by(SurveyPeriod.start_time.asc())
    )
    return list(db.scalars(stmt).all())


@app.post("/survey-periods", response_model=SurveyPeriodOut)
def create_survey_period(
    payload: SurveyPeriodUpsertIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> SurveyPeriodOut:
    _validate_period_payload(db, payload)

    period = SurveyPeriod(
        period_code=payload.period_code,
        period_name=payload.period_name,
        start_time=payload.start_time,
        end_time=payload.end_time,
        period_type=payload.period_type,
        month_no=payload.month_no,
        half_no=payload.half_no,
        status=payload.status,
        created_by_user_id=current_user.id,
        updated_by_user_id=current_user.id,
    )
    db.add(period)
    db.flush()

    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="period_create",
        target_type="survey_period",
        target_id=str(period.id),
        reason=None,
        before_state=None,
        after_state={
            "period_code": period.period_code,
            "period_name": period.period_name,
            "period_type": period.period_type.value,
            "month_no": period.month_no,
            "half_no": period.half_no,
            "status": period.status.value,
        },
    )

    db.commit()
    db.refresh(period)
    return period


@app.put("/survey-periods/{period_id}", response_model=SurveyPeriodOut)
def update_survey_period(
    period_id: int,
    payload: SurveyPeriodUpsertIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> SurveyPeriodOut:
    period = db.get(SurveyPeriod, period_id)
    if period is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="调查期不存在")

    _validate_period_payload(db, payload, exclude_period_id=period.id)

    before_state = {
        "period_code": period.period_code,
        "period_name": period.period_name,
        "period_type": period.period_type.value,
        "month_no": period.month_no,
        "half_no": period.half_no,
        "status": period.status.value,
    }

    period.period_code = payload.period_code
    period.period_name = payload.period_name
    period.start_time = payload.start_time
    period.end_time = payload.end_time
    period.period_type = payload.period_type
    period.month_no = payload.month_no
    period.half_no = payload.half_no
    period.status = payload.status
    period.updated_by_user_id = current_user.id

    db.flush()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="period_update",
        target_type="survey_period",
        target_id=str(period.id),
        reason=None,
        before_state=before_state,
        after_state={
            "period_code": period.period_code,
            "period_name": period.period_name,
            "period_type": period.period_type.value,
            "month_no": period.month_no,
            "half_no": period.half_no,
            "status": period.status.value,
        },
    )

    db.commit()
    db.refresh(period)
    return period


@app.get("/filings/me", response_model=FilingOut)
def get_my_filing(
    current_user: User = Depends(require_roles(UserRole.ENTERPRISE)),
    db: Session = Depends(get_db),
) -> FilingOut:
    enterprise = db.get(Enterprise, current_user.enterprise_id) if current_user.enterprise_id else None
    if enterprise is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="未找到企业备案信息")
    return enterprise


@app.post("/filings/submit", response_model=FilingOut)
def submit_filing(
    payload: FilingPayload,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.ENTERPRISE)),
    db: Session = Depends(get_db),
) -> FilingOut:
    enterprise = _get_or_create_enterprise(db, current_user)

    if (
        enterprise.filing_status == FilingStatus.APPROVED
        and enterprise.organization_code
        and enterprise.organization_code != payload.organization_code
    ):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="已备案企业不可修改组织机构代码")

    duplicate = db.scalar(
        select(Enterprise).where(
            Enterprise.organization_code == payload.organization_code,
            Enterprise.id != enterprise.id,
        )
    )
    if duplicate is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="组织机构代码已存在")

    before_state = _enterprise_snapshot(enterprise)

    enterprise.organization_code = payload.organization_code
    enterprise.name = payload.name
    enterprise.nature = payload.nature
    enterprise.industry = payload.industry
    enterprise.business_scope = payload.business_scope
    enterprise.contact_person = payload.contact_person
    enterprise.contact_address = payload.contact_address
    enterprise.postal_code = payload.postal_code
    enterprise.phone = payload.phone
    enterprise.fax = payload.fax
    enterprise.email = payload.email
    if not enterprise.city_code:
        enterprise.city_code = current_user.city_code or "530100"
    if not enterprise.city_name:
        enterprise.city_name = "昆明市"

    enterprise.filing_status = FilingStatus.PENDING
    enterprise.filing_reject_reason = None
    enterprise.filing_submit_time = datetime.utcnow()

    db.flush()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="filing_submit",
        target_type="enterprise_filing",
        target_id=str(enterprise.id),
        reason=None,
        before_state=before_state,
        after_state=_enterprise_snapshot(enterprise),
    )

    db.commit()
    db.refresh(enterprise)
    return enterprise


@app.post("/filings/save", response_model=FilingOut)
def save_filing_draft(
    payload: FilingPayload,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.ENTERPRISE)),
    db: Session = Depends(get_db),
) -> FilingOut:
    enterprise = _get_or_create_enterprise(db, current_user)

    duplicate = db.scalar(
        select(Enterprise).where(
            Enterprise.organization_code == payload.organization_code,
            Enterprise.id != enterprise.id,
        )
    )
    if duplicate is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="组织机构代码已存在")

    before_state = _enterprise_snapshot(enterprise)

    enterprise.organization_code = payload.organization_code
    enterprise.name = payload.name
    enterprise.nature = payload.nature
    enterprise.industry = payload.industry
    enterprise.business_scope = payload.business_scope
    enterprise.contact_person = payload.contact_person
    enterprise.contact_address = payload.contact_address
    enterprise.postal_code = payload.postal_code
    enterprise.phone = payload.phone
    enterprise.fax = payload.fax
    enterprise.email = payload.email
    if not enterprise.city_code:
        enterprise.city_code = current_user.city_code or "530100"
    if not enterprise.city_name:
        enterprise.city_name = "昆明市"

    if enterprise.filing_status == FilingStatus.NOT_SUBMITTED:
        enterprise.filing_status = FilingStatus.NOT_SUBMITTED

    db.flush()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="filing_save",
        target_type="enterprise_filing",
        target_id=str(enterprise.id),
        reason=None,
        before_state=before_state,
        after_state=_enterprise_snapshot(enterprise),
    )

    db.commit()
    db.refresh(enterprise)
    return enterprise


def _query_filings(
    db: Session,
    filing_status_value: FilingStatus | None,
    name: str | None,
    region: str | None,
    period_code: str | None,
) -> list[Enterprise]:
    stmt = select(Enterprise).order_by(Enterprise.updated_at.desc())
    if period_code:
        stmt = (
            stmt.join(Enterprise.reports)
            .where(EmploymentReport.period_code == period_code, _report_is_active_filter())
            .distinct()
        )
    if filing_status_value is not None:
        stmt = stmt.where(Enterprise.filing_status == filing_status_value)
    if name:
        stmt = stmt.where(Enterprise.name.contains(name.strip()))
    if region:
        region_q = region.strip()
        stmt = stmt.where(or_(Enterprise.city_name.contains(region_q), Enterprise.city_code.contains(region_q)))
    return list(db.scalars(stmt.limit(500)).all())


@app.get("/filings", response_model=list[FilingOut])
def list_filings(
    filing_status_value: FilingStatus | None = None,
    name: str | None = None,
    region: str | None = None,
    period_code: str | None = None,
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> list[FilingOut]:
    return _query_filings(db, filing_status_value, name, region, period_code)


@app.get("/province/filings/export/xlsx")
def province_filing_export_xlsx(
    filing_status_value: FilingStatus | None = None,
    name: str | None = None,
    region: str | None = None,
    period_code: str | None = None,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> Response:
    filings = _query_filings(db, filing_status_value, name, region, period_code)
    rows = [
        [
            item.id,
            item.name or "",
            item.organization_code or "",
            item.city_name or item.city_code or "",
            item.contact_person or "",
            item.phone or "",
            item.filing_status.value,
            item.updated_at,
        ]
        for item in filings
    ]
    return _build_xlsx_response(
        title="省级企业备案列表导出",
        headers=["企业ID", "企业名称", "组织机构代码", "所属地市", "联系人", "联系电话", "备案状态", "更新时间"],
        rows=rows,
        query_filters={"备案状态": filing_status_value, "企业名称": name, "地区": region, "调查期": period_code},
        operator=current_user,
        file_name=f"province_filing_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )


@app.get("/filings/{enterprise_id}", response_model=FilingOut)
def filing_detail(
    enterprise_id: int,
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> FilingOut:
    enterprise = db.get(Enterprise, enterprise_id)
    if enterprise is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="备案信息不存在")
    return enterprise


@app.post("/filings/{enterprise_id}/approve", response_model=FilingOut)
def approve_filing(
    enterprise_id: int,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> FilingOut:
    enterprise = db.get(Enterprise, enterprise_id)
    if enterprise is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="备案信息不存在")
    if enterprise.filing_status != FilingStatus.PENDING:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="仅待备案状态可审批通过")

    before_state = _enterprise_snapshot(enterprise)

    enterprise.filing_status = FilingStatus.APPROVED
    enterprise.filing_reject_reason = None
    enterprise.filing_review_time = datetime.utcnow()

    db.flush()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="filing_approve",
        target_type="enterprise_filing",
        target_id=str(enterprise.id),
        reason=None,
        before_state=before_state,
        after_state=_enterprise_snapshot(enterprise),
    )

    db.commit()
    db.refresh(enterprise)
    return enterprise


@app.post("/filings/{enterprise_id}/reject", response_model=FilingOut)
def reject_filing(
    enterprise_id: int,
    payload: FilingReviewIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> FilingOut:
    enterprise = db.get(Enterprise, enterprise_id)
    if enterprise is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="备案信息不存在")
    if enterprise.filing_status != FilingStatus.PENDING:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="仅待备案状态可退回")

    reason = (payload.reason or "").strip()
    if not reason:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="退回必须填写原因")

    before_state = _enterprise_snapshot(enterprise)

    enterprise.filing_status = FilingStatus.REJECTED
    enterprise.filing_reject_reason = reason
    enterprise.filing_review_time = datetime.utcnow()

    db.flush()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="filing_reject",
        target_type="enterprise_filing",
        target_id=str(enterprise.id),
        reason=reason,
        before_state=before_state,
        after_state=_enterprise_snapshot(enterprise),
    )

    db.commit()
    db.refresh(enterprise)
    return enterprise


@app.get("/reports/me", response_model=list[ReportOut])
def my_reports(
    period_code: str | None = None,
    report_status: ReportStatus | None = None,
    current_user: User = Depends(require_roles(UserRole.ENTERPRISE)),
    db: Session = Depends(get_db),
) -> list[ReportOut]:
    enterprise = db.get(Enterprise, current_user.enterprise_id) if current_user.enterprise_id else None
    if enterprise is None:
        return []

    stmt = (
        select(EmploymentReport)
        .options(joinedload(EmploymentReport.survey_period), joinedload(EmploymentReport.enterprise))
        .where(EmploymentReport.enterprise_id == enterprise.id, _report_is_active_filter())
        .order_by(EmploymentReport.updated_at.desc())
    )
    if period_code is not None:
        stmt = stmt.where(EmploymentReport.period_code == period_code)
    if report_status is not None:
        stmt = stmt.where(EmploymentReport.status == report_status)

    reports = list(db.scalars(stmt).all())
    return [_report_to_out(item) for item in reports]


@app.get("/reports/available-periods", response_model=list[SurveyPeriodOut])
def report_available_periods(
    current_user: User = Depends(require_roles(UserRole.ENTERPRISE)),
    db: Session = Depends(get_db),
) -> list[SurveyPeriodOut]:
    enterprise = db.get(Enterprise, current_user.enterprise_id) if current_user.enterprise_id else None
    if enterprise is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="企业信息不存在")
    if enterprise.filing_status != FilingStatus.APPROVED:
        if enterprise.filing_status == FilingStatus.PENDING:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="备案已提交，待省级审核通过后才能填报")
        if enterprise.filing_status == FilingStatus.REJECTED:
            reason = enterprise.filing_reject_reason or "请修改备案信息后重新提交"
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"备案已退回：{reason}")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请先提交备案并通过省级审核后再填报")

    now = datetime.utcnow()
    stmt = (
        select(SurveyPeriod)
        .where(
            SurveyPeriod.status == SurveyPeriodStatus.ENABLED,
            SurveyPeriod.start_time <= now,
            SurveyPeriod.end_time >= now,
        )
        .order_by(SurveyPeriod.start_time.asc())
    )
    return list(db.scalars(stmt).all())


@app.get("/reports/current", response_model=ReportOut | None)
def current_period_report(
    period_code: str,
    current_user: User = Depends(require_roles(UserRole.ENTERPRISE)),
    db: Session = Depends(get_db),
) -> ReportOut | None:
    enterprise = db.get(Enterprise, current_user.enterprise_id) if current_user.enterprise_id else None
    if enterprise is None:
        return None

    period = _get_period_by_code(db, period_code)
    stmt = (
        select(EmploymentReport)
        .options(joinedload(EmploymentReport.survey_period), joinedload(EmploymentReport.enterprise))
        .where(
            EmploymentReport.enterprise_id == enterprise.id,
            EmploymentReport.survey_period_id == period.id,
            _report_is_active_filter(),
        )
    )
    report = db.scalar(stmt)
    if report is None:
        return None
    return _report_to_out(report)


@app.post("/reports/save", response_model=ReportOut)
def save_report(
    payload: ReportPayload,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.ENTERPRISE)),
    db: Session = Depends(get_db),
) -> ReportOut:
    enterprise = _get_or_create_enterprise(db, current_user)
    report, before_state = _upsert_report_for_enterprise(db, enterprise, payload)

    report.status = ReportStatus.DRAFT
    report.updated_by_user_id = current_user.id

    db.flush()
    _save_report_version(db, report, current_user, "report_save")
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="report_save",
        target_type="employment_report",
        target_id=str(report.id),
        reason=None,
        before_state=before_state,
        after_state=_report_snapshot(report),
    )

    db.commit()
    report = _get_report_with_relations(db, report.id)
    return _report_to_out(report)


@app.post("/reports/submit", response_model=ReportOut)
def submit_report(
    payload: ReportPayload,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.ENTERPRISE)),
    db: Session = Depends(get_db),
) -> ReportOut:
    enterprise = _get_or_create_enterprise(db, current_user)
    _validate_submission(payload)

    report, before_state = _upsert_report_for_enterprise(db, enterprise, payload)
    report.status = ReportStatus.PENDING_CITY_REVIEW
    report.updated_by_user_id = current_user.id
    report.last_submitted_at = datetime.utcnow()

    db.flush()
    _save_report_version(db, report, current_user, "report_submit")
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="report_submit",
        target_type="employment_report",
        target_id=str(report.id),
        reason=None,
        before_state=before_state,
        after_state=_report_snapshot(report),
    )

    db.commit()
    report = _get_report_with_relations(db, report.id)
    return _report_to_out(report)


@app.get("/city/reports", response_model=list[ReportOut])
def list_city_reports(
    report_status: ReportStatus | None = None,
    current_user: User = Depends(require_roles(UserRole.CITY)),
    db: Session = Depends(get_db),
) -> list[ReportOut]:
    return _list_city_reports(db, current_user, report_status)


@app.get("/city/reports/{report_id}", response_model=ReportOut)
def city_report_detail(
    report_id: int,
    current_user: User = Depends(require_roles(UserRole.CITY)),
    db: Session = Depends(get_db),
) -> ReportOut:
    report = _get_report_with_relations(db, report_id)
    _assert_city_scope(current_user, report)
    return _report_to_out(report)


@app.post("/city/reports/{report_id}/approve", response_model=ReportOut)
def city_approve_report(
    report_id: int,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.CITY)),
    db: Session = Depends(get_db),
) -> ReportOut:
    report = _get_report_with_relations(db, report_id)
    _assert_city_scope(current_user, report)
    if report.status != ReportStatus.PENDING_CITY_REVIEW:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="仅待市审核报表可审核通过")

    before_state = _report_snapshot(report)

    report.status = ReportStatus.CITY_APPROVED
    report.city_reviewed_at = datetime.utcnow()
    report.updated_by_user_id = current_user.id

    db.flush()
    _save_report_version(db, report, current_user, "report_city_approve")
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="report_city_approve",
        target_type="employment_report",
        target_id=str(report.id),
        reason=None,
        before_state=before_state,
        after_state=_report_snapshot(report),
    )

    db.commit()
    report = _get_report_with_relations(db, report.id)
    return _report_to_out(report)


@app.post("/city/reports/{report_id}/reject", response_model=ReportOut)
def city_reject_report(
    report_id: int,
    payload: ReportReviewIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.CITY)),
    db: Session = Depends(get_db),
) -> ReportOut:
    report = _get_report_with_relations(db, report_id)
    _assert_city_scope(current_user, report)
    if report.status != ReportStatus.PENDING_CITY_REVIEW:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="仅待市审核报表可退回")

    reason = payload.reason.strip()
    if not reason:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="退回必须填写原因")

    before_state = _report_snapshot(report)

    report.status = ReportStatus.CITY_REJECTED
    report.city_reviewed_at = datetime.utcnow()
    report.updated_by_user_id = current_user.id

    db.flush()
    _save_report_version(db, report, current_user, "report_city_reject")
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="report_city_reject",
        target_type="employment_report",
        target_id=str(report.id),
        reason=reason,
        before_state=before_state,
        after_state=_report_snapshot(report),
    )

    db.commit()
    report = _get_report_with_relations(db, report.id)
    return _report_to_out(report)


@app.get("/province/reports", response_model=list[ReportOut])
def list_province_reports(
    report_status: ReportStatus | None = None,
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> list[ReportOut]:
    return _list_province_reports(db, report_status)


@app.get("/province/reports/{report_id}", response_model=ReportOut)
def province_report_detail(
    report_id: int,
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> ReportOut:
    report = _get_report_with_relations(db, report_id)
    return _report_to_out(report)


@app.post("/province/reports/{report_id}/approve", response_model=ReportOut)
def province_approve_report(
    report_id: int,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> ReportOut:
    report = _get_report_with_relations(db, report_id)
    if report.status != ReportStatus.CITY_APPROVED:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="仅待省审核报表可审核通过")

    before_state = _report_snapshot(report)

    report.status = ReportStatus.PROVINCE_APPROVED
    report.province_reviewed_at = datetime.utcnow()
    report.updated_by_user_id = current_user.id

    db.flush()
    _save_report_version(db, report, current_user, "report_province_approve")
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="report_province_approve",
        target_type="employment_report",
        target_id=str(report.id),
        reason=None,
        before_state=before_state,
        after_state=_report_snapshot(report),
    )

    db.commit()
    report = _get_report_with_relations(db, report.id)
    return _report_to_out(report)


@app.post("/province/reports/{report_id}/reject", response_model=ReportOut)
def province_reject_report(
    report_id: int,
    payload: ReportReviewIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> ReportOut:
    report = _get_report_with_relations(db, report_id)
    if report.status != ReportStatus.CITY_APPROVED:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="仅待省审核报表可退回")

    reason = payload.reason.strip()
    if not reason:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="退回必须填写原因")

    before_state = _report_snapshot(report)

    report.status = ReportStatus.PROVINCE_REJECTED
    report.province_reviewed_at = datetime.utcnow()
    report.updated_by_user_id = current_user.id

    db.flush()
    _save_report_version(db, report, current_user, "report_province_reject")
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="report_province_reject",
        target_type="employment_report",
        target_id=str(report.id),
        reason=reason,
        before_state=before_state,
        after_state=_report_snapshot(report),
    )

    db.commit()
    report = _get_report_with_relations(db, report.id)
    return _report_to_out(report)


@app.post("/province/reports/{report_id}/submit", response_model=ReportOut)
def province_submit_report(
    report_id: int,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> ReportOut:
    report = _get_report_with_relations(db, report_id)
    if report.status != ReportStatus.PROVINCE_APPROVED:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="仅省审核通过报表可执行上报")

    before_state = _report_snapshot(report)

    report.status = ReportStatus.PROVINCE_SUBMITTED
    report.updated_by_user_id = current_user.id

    db.flush()
    _save_report_version(db, report, current_user, "report_province_submit")
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="report_province_submit",
        target_type="employment_report",
        target_id=str(report.id),
        reason=None,
        before_state=before_state,
        after_state=_report_snapshot(report),
    )

    db.commit()
    report = _get_report_with_relations(db, report.id)
    return _report_to_out(report)


@app.post("/province/reports/{report_id}/modify", response_model=ReportOut)
def province_modify_report(
    report_id: int,
    payload: ReportModifyIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> ReportOut:
    report = _get_report_with_relations(db, report_id)

    reason = payload.reason.strip()
    if not reason:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="修改原因必填")

    validate_payload = ReportPayload(
        period_code=report.period_code,
        base_employment=payload.base_employment,
        survey_employment=payload.survey_employment,
        decrease_type=payload.decrease_type,
        decrease_reason=payload.decrease_reason,
        decrease_reason_detail=payload.decrease_reason_detail,
        other_note=payload.other_note,
    )
    _validate_submission(validate_payload)

    before_state = _report_snapshot(report)

    report.base_employment = payload.base_employment
    report.survey_employment = payload.survey_employment
    report.decrease_type = payload.decrease_type
    report.decrease_reason = payload.decrease_reason
    report.decrease_reason_detail = payload.decrease_reason_detail
    report.other_note = payload.other_note
    report.updated_by_user_id = current_user.id

    if report.survey_employment >= report.base_employment:
        report.decrease_type = None
        report.decrease_reason = None
        report.decrease_reason_detail = None
        report.other_note = None

    db.flush()
    _save_report_version(db, report, current_user, "report_province_modify")
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="report_province_modify",
        target_type="employment_report",
        target_id=str(report.id),
        reason=reason,
        before_state=before_state,
        after_state=_report_snapshot(report),
    )

    db.commit()
    report = _get_report_with_relations(db, report.id)
    return _report_to_out(report)


@app.get("/province/reports/{report_id}/delete-check")
def province_report_delete_check(
    report_id: int,
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    report = _get_report_with_relations(db, report_id)
    allowed, reason = _check_delete_allowed(report)
    return {"allowed": allowed, "reason": reason}


@app.delete("/province/reports/{report_id}")
def province_delete_report(
    report_id: int,
    payload: ReportReviewIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    report = _get_report_with_relations(db, report_id)
    if report.is_deleted:
        return {"deleted": True, "report_id": report.id}

    allowed, blocked_reason = _check_delete_allowed(report)
    if not allowed:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=blocked_reason or "当前报表禁止删除")

    reason = payload.reason.strip()
    if not reason:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="删除原因必填")

    before_state = _report_snapshot(report)
    report.is_deleted = True
    report.deleted_at = datetime.utcnow()
    report.updated_by_user_id = current_user.id
    db.flush()

    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="report_delete",
        target_type="employment_report",
        target_id=str(report.id),
        reason=reason,
        before_state=before_state,
        after_state={"is_deleted": True},
    )
    db.commit()
    return {"deleted": True, "report_id": report.id}


@app.get("/province/summary", response_model=ProvinceSummaryOut)
def province_summary(
    period_code: str | None = None,
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> ProvinceSummaryOut:
    period = _resolve_analytic_period(db, period_code)
    by_city = _province_city_summary_rows(db, period.id)
    enterprise_total = sum(item.enterprise_count for item in by_city)
    base_total = sum(item.base_total for item in by_city)
    survey_total = sum(item.survey_total for item in by_city)
    return ProvinceSummaryOut(
        period_code=period.period_code,
        enterprise_total=enterprise_total,
        base_total=base_total,
        survey_total=survey_total,
        change_total=survey_total - base_total,
        by_city=by_city,
    )


@app.get("/province/analysis/compare", response_model=ProvinceCompareOut)
def province_analysis_compare(
    period_code_a: str,
    period_code_b: str,
    dimension: str = "region",
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> ProvinceCompareOut:
    period_a = _get_period_by_code(db, period_code_a)
    period_b = _get_period_by_code(db, period_code_b)

    rows_a_raw = _province_dimension_summary_rows(db, period_a.id, dimension)
    rows_b_raw = _province_dimension_summary_rows(db, period_b.id, dimension)
    rows_a = {item["value"]: item for item in rows_a_raw}
    rows_b = {item["value"]: item for item in rows_b_raw}

    all_cities = sorted(set(rows_a.keys()) | set(rows_b.keys()))
    by_city: list[CompareCityOut] = []
    for city in all_cities:
        item_a = rows_a.get(city)
        item_b = rows_b.get(city)
        by_city.append(
            CompareCityOut(
                city=city,
                base_a=item_a["base_total"] if item_a else 0,
                base_b=item_b["base_total"] if item_b else 0,
                survey_a=item_a["survey_total"] if item_a else 0,
                survey_b=item_b["survey_total"] if item_b else 0,
                report_count_a=item_a["enterprise_count"] if item_a else 0,
                report_count_b=item_b["enterprise_count"] if item_b else 0,
            )
        )

    enterprise_a = sum(item.report_count_a for item in by_city)
    enterprise_b = sum(item.report_count_b for item in by_city)
    base_a = sum(item.base_a for item in by_city)
    base_b = sum(item.base_b for item in by_city)
    survey_a = sum(item.survey_a for item in by_city)
    survey_b = sum(item.survey_b for item in by_city)
    change_a = survey_a - base_a
    change_b = survey_b - base_b
    decrease_a = sum(item.get("decrease_total", 0) for item in rows_a_raw)
    decrease_b = sum(item.get("decrease_total", 0) for item in rows_b_raw)
    change_ratio_a = round((change_a / base_a) * 100, 2) if base_a else 0.0
    change_ratio_b = round((change_b / base_b) * 100, 2) if base_b else 0.0

    metrics = [
        CompareMetricOut(metric="企业总数", value_a=float(enterprise_a), value_b=float(enterprise_b), ratio=_ratio_text(enterprise_a, enterprise_b)),
        CompareMetricOut(metric="建档期总岗位数", value_a=float(base_a), value_b=float(base_b), ratio=_ratio_text(base_a, base_b)),
        CompareMetricOut(metric="调查期总岗位数", value_a=float(survey_a), value_b=float(survey_b), ratio=_ratio_text(survey_a, survey_b)),
        CompareMetricOut(metric="岗位变化总数", value_a=float(change_a), value_b=float(change_b), ratio=_ratio_text(change_a, change_b)),
        CompareMetricOut(metric="岗位减少总数", value_a=float(decrease_a), value_b=float(decrease_b), ratio=_ratio_text(decrease_a, decrease_b)),
        CompareMetricOut(metric="岗位变化数量占比(%)", value_a=change_ratio_a, value_b=change_ratio_b, ratio=_ratio_text(int(change_ratio_a), int(change_ratio_b))),
    ]

    return ProvinceCompareOut(
        period_code_a=period_a.period_code,
        period_code_b=period_b.period_code,
        by_city=by_city,
        metrics=metrics,
    )


@app.get("/province/analysis/trend", response_model=ProvinceTrendOut)
def province_analysis_trend(
    period_codes: str | None = None,
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> ProvinceTrendOut:
    code_list = [item.strip() for item in period_codes.split(",")] if period_codes else []
    code_list = [item for item in code_list if item]

    stmt = (
        select(
            SurveyPeriod.period_code,
            SurveyPeriod.period_name,
            SurveyPeriod.start_time,
            func.coalesce(func.sum(EmploymentReport.base_employment), 0).label("base_total"),
            func.coalesce(func.sum(EmploymentReport.survey_employment), 0).label("survey_total"),
        )
        .join(EmploymentReport, EmploymentReport.survey_period_id == SurveyPeriod.id)
        .where(EmploymentReport.status.in_(PROVINCE_ANALYTIC_STATUSES), _report_is_active_filter())
    )
    if code_list:
        stmt = stmt.where(SurveyPeriod.period_code.in_(code_list))

    stmt = stmt.group_by(SurveyPeriod.period_code, SurveyPeriod.period_name, SurveyPeriod.start_time).order_by(
        SurveyPeriod.start_time.asc()
    )

    rows = list(db.execute(stmt).all())
    if not code_list and len(rows) > 6:
        rows = rows[-6:]

    points = [
        TrendPointOut(
            period_code=row.period_code,
            period_name=row.period_name,
            base_total=int(row.base_total or 0),
            survey_total=int(row.survey_total or 0),
            change_total=int((row.survey_total or 0) - (row.base_total or 0)),
            change_ratio=(
                round((((row.survey_total or 0) - (row.base_total or 0)) / (row.base_total or 1)) * 100, 2)
                if (row.base_total or 0) > 0
                else 0.0
            ),
        )
        for row in rows
    ]
    return ProvinceTrendOut(points=points)


@app.get("/province/analysis/multi-dim", response_model=ProvinceMultiDimOut)
def province_analysis_multi_dim(
    period_code: str | None = None,
    dimension: str = "region",
    city: str | None = None,
    nature: str | None = None,
    industry: str | None = None,
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> ProvinceMultiDimOut:
    period = _resolve_analytic_period(db, period_code)
    rows = _province_dimension_summary_rows(
        db,
        period.id,
        dimension=dimension,
        city=city,
        nature=nature,
        industry=industry,
    )
    return ProvinceMultiDimOut(
        period_code=period.period_code,
        dimension=dimension,
        items=[
            MultiDimItemOut(
                dimension=dimension,
                value=item["value"],
                enterprise_count=item["enterprise_count"],
                base_total=item["base_total"],
                survey_total=item["survey_total"],
                change_total=item["change_total"],
                decrease_total=item["decrease_total"],
                change_ratio=item["change_ratio"],
            )
            for item in rows
        ],
    )


@app.get("/province/sampling", response_model=ProvinceSamplingOut)
def province_sampling(
    period_code: str | None = None,
    city: str | None = None,
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> ProvinceSamplingOut:
    period = _resolve_analytic_period(db, period_code)
    by_city = _province_city_summary_rows(db, period.id)
    if city:
        by_city = [item for item in by_city if item.city == city]

    total_enterprises = sum(item.enterprise_count for item in by_city)
    sampling_rows = [
        SamplingCityOut(
            city=item.city,
            enterprise_count=item.enterprise_count,
            ratio=round((item.enterprise_count / total_enterprises), 4) if total_enterprises else 0,
        )
        for item in by_city
    ]
    return ProvinceSamplingOut(
        period_code=period.period_code,
        total_enterprises=total_enterprises,
        city_count=len(sampling_rows),
        by_city=sampling_rows,
    )


@app.get("/reports", response_model=list[ReportOut])
def list_reports(
    report_status: ReportStatus | None = None,
    current_user: User = Depends(require_roles(UserRole.CITY, UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> list[ReportOut]:
    if current_user.role == UserRole.CITY:
        return _list_city_reports(db, current_user, report_status)
    return _list_province_reports(db, report_status)


@app.get("/reports/{report_id}", response_model=ReportOut)
def report_detail(
    report_id: int,
    current_user: User = Depends(require_roles(UserRole.CITY, UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> ReportOut:
    report = _get_report_with_relations(db, report_id)
    if current_user.role == UserRole.CITY:
        _assert_city_scope(current_user, report)
    return _report_to_out(report)


@app.post("/reports/{report_id}/city-reject", response_model=ReportOut)
def city_reject_report_compat(
    report_id: int,
    payload: ReportReviewIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.CITY)),
    db: Session = Depends(get_db),
) -> ReportOut:
    return city_reject_report(report_id, payload, request, current_user, db)


@app.get("/reports/{report_id}/versions", response_model=list[ReportVersionOut])
def list_report_versions(
    report_id: int,
    current_user: User = Depends(require_roles(UserRole.ENTERPRISE, UserRole.CITY, UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> list[ReportVersionOut]:
    report = _get_report_with_relations(db, report_id)

    if current_user.role == UserRole.ENTERPRISE:
        if report.enterprise_id != current_user.enterprise_id:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="仅可查看本企业报表版本")
    elif current_user.role == UserRole.CITY:
        _assert_city_scope(current_user, report)

    stmt = (
        select(ReportVersion)
        .where(ReportVersion.report_id == report_id)
        .order_by(ReportVersion.version_no.desc())
    )
    return list(db.scalars(stmt).all())


@app.get("/audit-logs", response_model=list[AuditLogOut])
def list_audit_logs(
    action_type: str | None = None,
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> list[AuditLogOut]:
    stmt = select(AuditLog).order_by(AuditLog.created_at.desc())
    if action_type:
        stmt = stmt.where(AuditLog.action_type == action_type)
    stmt = stmt.limit(200)
    return list(db.scalars(stmt).all())


@app.get("/operation-logs", response_model=list[OperationLogOut])
def list_operation_logs(
    operation_type: str | None = None,
    user_name: str | None = None,
    target_type: str | None = None,
    keyword: str | None = None,
    start_time: datetime | None = None,
    end_time: datetime | None = None,
    limit: int = 300,
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> list[OperationLogOut]:
    stmt = select(OperationLog).order_by(OperationLog.operation_time.desc())
    if operation_type:
        stmt = stmt.where(OperationLog.operation_type == operation_type)
    if user_name:
        stmt = stmt.where(OperationLog.user_name.contains(user_name.strip()))
    if target_type:
        stmt = stmt.where(OperationLog.target_type == target_type)
    if keyword:
        key = keyword.strip()
        stmt = stmt.where(
            or_(
                OperationLog.target_id.contains(key),
                OperationLog.reason.contains(key),
                OperationLog.operation_type.contains(key),
            )
        )
    if start_time is not None:
        stmt = stmt.where(OperationLog.operation_time >= start_time)
    if end_time is not None:
        stmt = stmt.where(OperationLog.operation_time <= end_time)
    safe_limit = min(max(limit, 1), 1000)
    stmt = stmt.limit(safe_limit)
    return list(db.scalars(stmt).all())


def _query_export_rows(
    db: Session,
    unit_name: str | None,
    account: str | None,
    user_type: str | None,
    city: str | None,
    county: str | None,
    area: str | None,
    report_status: ReportStatus | None,
    unit_nature: str | None,
    industry: str | None,
    period_code: str | None,
    stat_month: int | None,
    stat_quarter: int | None,
    start_time: datetime | None,
    end_time: datetime | None,
) -> list[ReportExportRowOut]:
    def split_address(value: str | None) -> tuple[str | None, str | None]:
        if not value:
            return None, None
        for sep in ["/", "-", "|", "，", ",", ">"]:
            if sep in value:
                parts = [item.strip() for item in value.split(sep) if item.strip()]
                if len(parts) >= 2:
                    return parts[0], parts[1]
                if len(parts) == 1:
                    return parts[0], None
        return value.strip(), None

    stmt = (
        select(EmploymentReport)
        .join(EmploymentReport.enterprise)
        .join(EmploymentReport.survey_period)
        .options(joinedload(EmploymentReport.enterprise), joinedload(EmploymentReport.survey_period))
        .where(_report_is_active_filter())
        .order_by(EmploymentReport.updated_at.desc())
    )
    if unit_name:
        stmt = stmt.where(Enterprise.name.contains(unit_name.strip()))
    if city:
        stmt = stmt.where(or_(Enterprise.city_name.contains(city.strip()), Enterprise.city_code.contains(city.strip())))
    if county:
        stmt = stmt.where(Enterprise.contact_address.contains(county.strip()))
    if area:
        stmt = stmt.where(Enterprise.contact_address.contains(area.strip()))
    if report_status is not None:
        stmt = stmt.where(EmploymentReport.status == report_status)
    if unit_nature:
        stmt = stmt.where(Enterprise.nature.contains(unit_nature.strip()))
    if industry:
        stmt = stmt.where(Enterprise.industry.contains(industry.strip()))
    if period_code:
        stmt = stmt.where(EmploymentReport.period_code == period_code)
    if stat_month is not None:
        stmt = stmt.where(SurveyPeriod.month_no == stat_month)
    if start_time is not None:
        stmt = stmt.where(EmploymentReport.updated_at >= start_time)
    if end_time is not None:
        stmt = stmt.where(EmploymentReport.updated_at <= end_time)

    reports = list(db.scalars(stmt.limit(500)).all())
    if not reports:
        return []

    enterprise_ids = {item.enterprise_id for item in reports}
    users = list(
        db.scalars(
            select(User)
            .where(User.role == UserRole.ENTERPRISE, User.enterprise_id.in_(enterprise_ids))
            .order_by(User.id.asc())
        ).all()
    )
    account_map: dict[int, tuple[str, str]] = {}
    for user in users:
        if user.enterprise_id is not None and user.enterprise_id not in account_map:
            account_map[user.enterprise_id] = (user.username, user.role.value)

    rows = [
        (
            lambda address_parts: ReportExportRowOut(
                report_id=item.id,
                unit_name=(item.enterprise.name if item.enterprise and item.enterprise.name else f"企业{item.enterprise_id}"),
                account=account_map.get(item.enterprise_id, (None, "enterprise"))[0],
                role="企业",
                user_type=account_map.get(item.enterprise_id, (None, "enterprise"))[1],
                city=item.enterprise.city_name if item.enterprise else None,
                county=address_parts[0],
                area=address_parts[1],
                unit_nature=item.enterprise.nature if item.enterprise else None,
                industry=item.enterprise.industry if item.enterprise else None,
                status=item.status,
                period_code=item.period_code,
                period_name=item.survey_period.period_name if item.survey_period else None,
                stat_month=item.survey_period.month_no if item.survey_period else None,
                stat_quarter=((item.survey_period.month_no - 1) // 3 + 1) if item.survey_period else None,
                base_employment=item.base_employment,
                survey_employment=item.survey_employment,
                updated_at=item.updated_at,
            )
        )(split_address(item.enterprise.contact_address if item.enterprise else None))
        for item in reports
    ]
    if account:
        account_q = account.strip()
        rows = [item for item in rows if item.account and account_q in item.account]
    if user_type:
        rows = [item for item in rows if item.user_type == user_type.strip()]
    if stat_quarter is not None:
        rows = [item for item in rows if item.stat_quarter == stat_quarter]
    return rows


@app.get("/province/export/query", response_model=list[ReportExportRowOut])
def province_export_query(
    unit_name: str | None = None,
    account: str | None = None,
    user_type: str | None = None,
    city: str | None = None,
    county: str | None = None,
    area: str | None = None,
    report_status: ReportStatus | None = None,
    unit_nature: str | None = None,
    industry: str | None = None,
    period_code: str | None = None,
    stat_month: int | None = None,
    stat_quarter: int | None = None,
    start_time: datetime | None = None,
    end_time: datetime | None = None,
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> list[ReportExportRowOut]:
    return _query_export_rows(
        db,
        unit_name,
        account,
        user_type,
        city,
        county,
        area,
        report_status,
        unit_nature,
        industry,
        period_code,
        stat_month,
        stat_quarter,
        start_time,
        end_time,
    )


@app.get("/province/export/csv")
def province_export_csv(
    unit_name: str | None = None,
    account: str | None = None,
    city: str | None = None,
    report_status: ReportStatus | None = None,
    period_code: str | None = None,
    start_time: datetime | None = None,
    end_time: datetime | None = None,
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> Response:
    rows = _query_export_rows(db, unit_name, account, city, report_status, period_code, start_time, end_time)
    lines = ["report_id,unit_name,account,role,city,status,period_code,period_name,base_employment,survey_employment,updated_at"]
    for item in rows:
        lines.append(
            f"{item.report_id},{item.unit_name},{item.account or ''},{item.role},{item.city or ''},{item.status.value},{item.period_code},{item.period_name or ''},{item.base_employment},{item.survey_employment},{item.updated_at.isoformat()}"
        )
    return Response(
        content="\n".join(lines),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=province_export.csv"},
    )


@app.get("/province/export/xlsx")
def province_export_xlsx(
    unit_name: str | None = None,
    account: str | None = None,
    user_type: str | None = None,
    city: str | None = None,
    county: str | None = None,
    area: str | None = None,
    report_status: ReportStatus | None = None,
    unit_nature: str | None = None,
    industry: str | None = None,
    period_code: str | None = None,
    stat_month: int | None = None,
    stat_quarter: int | None = None,
    start_time: datetime | None = None,
    end_time: datetime | None = None,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> Response:
    rows = _query_export_rows(
        db,
        unit_name,
        account,
        user_type,
        city,
        county,
        area,
        report_status,
        unit_nature,
        industry,
        period_code,
        stat_month,
        stat_quarter,
        start_time,
        end_time,
    )
    data_rows = [
        [
            item.report_id,
            item.unit_name,
            item.account or "",
            item.role,
            item.user_type,
            item.city or "",
            item.county or "",
            item.area or "",
            item.unit_nature or "",
            item.industry or "",
            item.status.value,
            item.period_code,
            item.period_name or "",
            item.stat_month or "",
            item.stat_quarter or "",
            item.base_employment,
            item.survey_employment,
            item.updated_at,
        ]
        for item in rows
    ]
    return _build_xlsx_response(
        title="省级数据查询导出",
        headers=[
            "报表ID",
            "单位名称",
            "登录账号",
            "用户类型",
            "角色编码",
            "所属地市",
            "所属市县",
            "所处区域",
            "单位性质",
            "所属行业",
            "数据状态",
            "调查期编码",
            "调查期名称",
            "统计月份",
            "统计季度",
            "建档期就业",
            "调查期就业",
            "更新时间",
        ],
        rows=data_rows,
        query_filters={
            "单位名称": unit_name,
            "账号": account,
            "用户类型": user_type,
            "地市": city,
            "市县": county,
            "区域": area,
            "状态": report_status,
            "单位性质": unit_nature,
            "所属行业": industry,
            "调查期": period_code,
            "统计月份": stat_month,
            "统计季度": stat_quarter,
            "开始时间": start_time,
            "结束时间": end_time,
        },
        operator=current_user,
        file_name=f"province_report_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )


@app.get("/system/users", response_model=list[SystemUserOut])
def list_system_users(
    role: UserRole | None = None,
    keyword: str | None = None,
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> list[SystemUserOut]:
    stmt = select(User).options(joinedload(User.custom_role)).order_by(User.created_at.desc())
    if role is not None:
        stmt = stmt.where(User.role == role)
    if keyword:
        stmt = stmt.where(User.username.contains(keyword.strip()))
    users = list(db.scalars(stmt.limit(300)).all())
    return [_system_user_to_out(item) for item in users]


@app.post("/system/users", response_model=SystemUserOut)
def create_system_user(
    payload: SystemUserCreateIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> SystemUserOut:
    exists = db.scalar(select(User).where(User.username == payload.username))
    if exists is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="用户名已存在")

    city_code = payload.city_code
    if payload.role == UserRole.CITY and not city_code:
        city_code = "530100"

    selected_custom_role = _resolve_custom_role(db, payload.custom_role_id, payload.role)

    user = User(
        username=payload.username,
        password_hash=hash_password(payload.password),
        role=payload.role,
        custom_role_id=selected_custom_role.id if selected_custom_role else None,
        is_active=payload.is_active,
        is_activated=payload.is_activated,
        city_code=city_code,
        enterprise_id=payload.enterprise_id,
    )
    db.add(user)
    db.flush()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="system_user_create",
        target_type="user",
        target_id=str(user.id),
        reason=None,
        before_state=None,
        after_state={
            "username": user.username,
            "role": user.role.value,
            "is_active": user.is_active,
            "is_activated": user.is_activated,
        },
    )
    db.commit()
    db.refresh(user)
    return _system_user_to_out(user)


@app.put("/system/users/{user_id}", response_model=SystemUserOut)
def update_system_user(
    user_id: int,
    payload: SystemUserUpdateIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> SystemUserOut:
    user = db.get(User, user_id)
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")

    before_state = {
        "role": user.role.value,
        "custom_role_id": user.custom_role_id,
        "is_active": user.is_active,
        "is_activated": user.is_activated,
        "city_code": user.city_code,
        "enterprise_id": user.enterprise_id,
    }
    if payload.role is not None:
        user.role = payload.role
        user.custom_role_id = None
    if "custom_role_id" in payload.model_fields_set:
        selected_custom_role = _resolve_custom_role(db, payload.custom_role_id, user.role)
        user.custom_role_id = selected_custom_role.id if selected_custom_role else None
    if payload.is_active is not None:
        user.is_active = payload.is_active
    if payload.is_activated is not None:
        user.is_activated = payload.is_activated
    if payload.city_code is not None:
        user.city_code = payload.city_code
    if payload.enterprise_id is not None:
        user.enterprise_id = payload.enterprise_id
    if payload.password:
        user.password_hash = hash_password(payload.password)

    if user.role == UserRole.CITY and not user.city_code:
        user.city_code = "530100"

    db.flush()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="system_user_update",
        target_type="user",
        target_id=str(user.id),
        reason=None,
        before_state=before_state,
        after_state={
            "role": user.role.value,
            "custom_role_id": user.custom_role_id,
            "is_active": user.is_active,
            "is_activated": user.is_activated,
            "city_code": user.city_code,
            "enterprise_id": user.enterprise_id,
        },
    )
    db.commit()
    db.refresh(user)
    return _system_user_to_out(user)


@app.get("/system/users/{user_id}/delete-check")
def system_user_delete_check(
    user_id: int,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    user = db.get(User, user_id)
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")
    allowed, reason = _check_user_delete_allowed(db, current_user, user)
    return {"allowed": allowed, "reason": reason}


@app.delete("/system/users/{user_id}")
def system_delete_user(
    user_id: int,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    user = db.get(User, user_id)
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")

    allowed, reason = _check_user_delete_allowed(db, current_user, user)
    if not allowed:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=reason or "当前用户禁止删除")

    before_state = {
        "username": user.username,
        "role": user.role.value,
        "is_active": user.is_active,
        "is_activated": user.is_activated,
    }
    user.is_active = False
    user.is_activated = False
    _clear_login_lock(user)
    db.flush()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="system_user_delete",
        target_type="user",
        target_id=str(user.id),
        reason="软删除为停用+未激活",
        before_state=before_state,
        after_state={
            "username": user.username,
            "is_active": user.is_active,
            "is_activated": user.is_activated,
        },
    )
    db.commit()
    return {"deleted": True, "user_id": user.id}


@app.get("/system/roles", response_model=list[SystemRoleOut])
def list_system_roles(
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.PROVINCE)),
) -> list[SystemRoleOut]:
    return _build_system_roles(db)


@app.post("/system/roles/custom", response_model=SystemRoleOut)
def create_system_custom_role(
    payload: SystemRoleCreateIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> SystemRoleOut:
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="角色名称不能为空")

    if name in {role.value for role in UserRole} or name in ROLE_NAME_MAP.values():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="角色名称与系统预定义角色冲突")

    exists = db.scalar(select(CustomRole).where(CustomRole.name == name))
    if exists is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="角色名称已存在")

    role_item = CustomRole(
        name=name,
        scope=payload.scope,
        permissions=_normalize_permissions(payload.permissions),
        is_builtin=False,
        created_by_user_id=current_user.id,
    )
    db.add(role_item)
    db.flush()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="system_role_create",
        target_type="custom_role",
        target_id=str(role_item.id),
        reason=None,
        before_state=None,
        after_state={"name": role_item.name, "scope": role_item.scope.value, "permissions": role_item.permissions},
    )
    db.commit()
    db.refresh(role_item)
    return _custom_role_to_out(db, role_item)


@app.put("/system/roles/custom/{role_id}", response_model=SystemRoleOut)
def update_system_custom_role(
    role_id: int,
    payload: SystemRoleUpdateIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> SystemRoleOut:
    role_item = db.get(CustomRole, role_id)
    if role_item is None or role_item.is_builtin:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="自定义角色不存在")

    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="角色名称不能为空")

    exists = db.scalar(select(CustomRole).where(CustomRole.name == name, CustomRole.id != role_item.id))
    if exists is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="角色名称已存在")

    before_state = {
        "name": role_item.name,
        "permissions": role_item.permissions,
    }
    role_item.name = name
    role_item.permissions = _normalize_permissions(payload.permissions)

    db.flush()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="system_role_update",
        target_type="custom_role",
        target_id=str(role_item.id),
        reason=None,
        before_state=before_state,
        after_state={"name": role_item.name, "permissions": role_item.permissions},
    )
    db.commit()
    db.refresh(role_item)
    return _custom_role_to_out(db, role_item)


@app.get("/system/roles/custom/{role_id}/delete-check", response_model=SystemRoleDeleteOut)
def system_custom_role_delete_check(
    role_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.PROVINCE)),
) -> SystemRoleDeleteOut:
    role_item = db.get(CustomRole, role_id)
    if role_item is None or role_item.is_builtin:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="自定义角色不存在")

    assigned_count = db.scalar(select(func.count(User.id)).where(User.custom_role_id == role_id)) or 0
    if assigned_count > 0:
        return SystemRoleDeleteOut(
            allowed=False,
            reason=f"该角色已分配给 {assigned_count} 个用户，删除前需先解除关联或选择强制解除",
            assigned_user_count=assigned_count,
        )
    return SystemRoleDeleteOut(allowed=True, reason=None, assigned_user_count=0)


@app.delete("/system/roles/custom/{role_id}")
def delete_system_custom_role(
    role_id: int,
    request: Request,
    detach_assigned_users: bool = False,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    role_item = db.get(CustomRole, role_id)
    if role_item is None or role_item.is_builtin:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="自定义角色不存在")

    assigned_count = db.scalar(select(func.count(User.id)).where(User.custom_role_id == role_id)) or 0
    if assigned_count > 0 and not detach_assigned_users:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"该角色已分配给 {assigned_count} 个用户，删除前需先解除关联",
        )

    if assigned_count > 0 and detach_assigned_users:
        db.execute(update(User).where(User.custom_role_id == role_id).values(custom_role_id=None))

    before_state = {
        "name": role_item.name,
        "scope": role_item.scope.value,
        "permissions": role_item.permissions,
        "assigned_user_count": assigned_count,
    }
    db.delete(role_item)
    db.flush()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="system_role_delete",
        target_type="custom_role",
        target_id=str(role_id),
        reason="删除自定义角色",
        before_state=before_state,
        after_state={"deleted": True, "detached_users": assigned_count if detach_assigned_users else 0},
    )
    db.commit()
    return {
        "deleted": True,
        "role_id": role_id,
        "detached_users": assigned_count if detach_assigned_users else 0,
    }


@app.get("/system/roles/{role_name}/delete-check")
def system_role_delete_check(
    role_name: UserRole,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.PROVINCE)),
) -> dict[str, Any]:
    user_count = db.scalar(select(func.count(User.id)).where(User.role == role_name)) or 0
    reason = "系统内置角色不可删除"
    if user_count > 0:
        reason += f"（当前绑定用户 {user_count} 人）"
    return {"allowed": False, "reason": reason}


@app.delete("/system/roles/{role_name}")
def system_delete_role(
    role_name: UserRole,
    _: User = Depends(require_roles(UserRole.PROVINCE)),
) -> dict[str, Any]:
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"角色 {role_name.value} 为系统内置角色，禁止删除")


@app.get("/system/monitor", response_model=SystemMonitorOut)
def system_monitor(
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> SystemMonitorOut:
    started = time.perf_counter()
    total_users = db.scalar(select(func.count(User.id))) or 0
    active_users = db.scalar(select(func.count(User.id)).where(User.is_active.is_(True))) or 0
    enterprises = db.scalar(select(func.count(Enterprise.id))) or 0
    reports = db.scalar(select(func.count(EmploymentReport.id)).where(_report_is_active_filter())) or 0
    notices = db.scalar(select(func.count(Notice.id)).where(Notice.is_deleted.is_(False))) or 0

    now = datetime.utcnow()
    active_period_ids = list(
        db.scalars(
            select(SurveyPeriod.id).where(
                SurveyPeriod.status == SurveyPeriodStatus.ENABLED,
                SurveyPeriod.start_time <= now,
                SurveyPeriod.end_time >= now,
            )
        ).all()
    )
    pending_review_count = 0
    if active_period_ids:
        pending_review_count = db.scalar(
            select(func.count(EmploymentReport.id)).where(
                EmploymentReport.survey_period_id.in_(active_period_ids),
                EmploymentReport.status.in_([ReportStatus.PENDING_CITY_REVIEW, ReportStatus.CITY_APPROVED]),
                _report_is_active_filter(),
            )
        ) or 0

    try:
        cpu_usage = int(psutil.cpu_percent(interval=0.0))
        memory_usage = int(psutil.virtual_memory().percent)
        disk_usage = int(psutil.disk_usage("/").percent)
    except Exception:
        cpu_usage = min(95, 35 + (active_users % 50))
        memory_usage = min(95, 40 + (reports % 45))
        disk_usage = min(95, 50 + (notices % 40))

    api_latency_ms = int((time.perf_counter() - started) * 1000)
    warning_level = _calc_warning_level(cpu_usage, memory_usage, disk_usage, pending_review_count, api_latency_ms)

    snapshot = SystemMonitorSnapshot(
        cpu_usage=cpu_usage,
        memory_usage=memory_usage,
        disk_usage=disk_usage,
        active_users=active_users,
        pending_review_count=pending_review_count,
        api_latency_ms=api_latency_ms,
        warning_level=warning_level,
    )
    db.add(snapshot)
    db.flush()

    stale = list(
        db.scalars(
            select(SystemMonitorSnapshot)
            .order_by(SystemMonitorSnapshot.created_at.desc())
            .offset(MONITOR_SNAPSHOT_KEEP)
        ).all()
    )
    for item in stale:
        db.delete(item)
    db.commit()

    return SystemMonitorOut(
        total_users=total_users,
        active_users=active_users,
        enterprises=enterprises,
        reports=reports,
        notices=notices,
        pending_review_count=pending_review_count,
        cpu_usage=cpu_usage,
        memory_usage=memory_usage,
        disk_usage=disk_usage,
        api_latency_ms=api_latency_ms,
        warning_level=warning_level,
    )


@app.get("/system/monitor/snapshots", response_model=list[SystemMonitorSnapshotOut])
def system_monitor_snapshots(
    limit: int = 10,
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> list[SystemMonitorSnapshotOut]:
    safe_limit = min(max(limit, 1), 50)
    stmt = select(SystemMonitorSnapshot).order_by(SystemMonitorSnapshot.created_at.desc()).limit(safe_limit)
    return list(db.scalars(stmt).all())


@app.post("/system/users/{user_id}/role-change")
def system_change_user_role(
    user_id: int,
    payload: SystemUserRoleChangeIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    user = db.get(User, user_id)
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")

    reason = payload.reason.strip()
    before_state = {"role": user.role.value}
    user.role = payload.new_role
    selected_custom_role = _resolve_custom_role(db, payload.custom_role_id, user.role)
    user.custom_role_id = selected_custom_role.id if selected_custom_role else None
    if user.role == UserRole.CITY and not user.city_code:
        user.city_code = "530100"

    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="user_role_change",
        target_type="user",
        target_id=str(user_id),
        reason=reason,
        before_state=before_state,
        after_state={"role": user.role.value, "custom_role_id": user.custom_role_id},
    )
    db.commit()
    return {
        "changed": True,
        "user_id": user.id,
        "role": user.role.value,
        "custom_role_id": user.custom_role_id,
    }


def _build_national_push_preview(db: Session, period_code: str | None) -> NationalPushPreviewOut:
    period = _resolve_analytic_period(db, period_code)
    reports = list(
        db.scalars(
            select(EmploymentReport)
            .options(joinedload(EmploymentReport.enterprise))
            .where(
                EmploymentReport.period_code == period.period_code,
                EmploymentReport.status == ReportStatus.PROVINCE_SUBMITTED,
                _report_is_active_filter(),
            )
        ).all()
    )
    payload = [
        {
            "report_id": item.id,
            "organization_code": item.enterprise.organization_code if item.enterprise else None,
            "enterprise_name": item.enterprise.name if item.enterprise else None,
            "city_code": item.enterprise.city_code if item.enterprise else None,
            "base_employment": item.base_employment,
            "survey_employment": item.survey_employment,
            "decrease_count": item.decrease_count,
        }
        for item in reports
    ]
    enterprise_count = len({item.enterprise_id for item in reports})
    return NationalPushPreviewOut(
        period_code=period.period_code,
        report_count=len(reports),
        enterprise_count=enterprise_count,
        payload=payload,
    )


@app.post("/national/push/preview", response_model=NationalPushPreviewOut)
def national_push_preview(
    period_code: str | None = None,
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> NationalPushPreviewOut:
    return _build_national_push_preview(db, period_code)


@app.get("/national/push/config", response_model=NationalPushApiConfigOut)
def national_push_config(
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> NationalPushApiConfigOut:
    config_row = db.scalar(select(NationalPushApiConfig).order_by(NationalPushApiConfig.updated_at.desc()))
    if config_row is None:
        return NationalPushApiConfigOut(enabled=False, api_key_masked="", updated_at=datetime.utcnow())
    return NationalPushApiConfigOut(
        enabled=config_row.enabled,
        api_key_masked=_mask_api_key(config_row.api_key),
        updated_at=config_row.updated_at,
    )


@app.post("/national/push/config", response_model=NationalPushApiConfigOut)
def upsert_national_push_config(
    payload: NationalPushApiConfigIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> NationalPushApiConfigOut:
    config_row = db.scalar(select(NationalPushApiConfig).order_by(NationalPushApiConfig.id.asc()))
    before_state = None
    if config_row is None:
        config_row = NationalPushApiConfig(
            api_key=payload.api_key.strip(),
            enabled=payload.enabled,
            updated_by_user_id=current_user.id,
        )
        db.add(config_row)
    else:
        before_state = {
            "enabled": config_row.enabled,
            "api_key_masked": _mask_api_key(config_row.api_key),
        }
        config_row.api_key = payload.api_key.strip()
        config_row.enabled = payload.enabled
        config_row.updated_by_user_id = current_user.id

    db.flush()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="national_push_config_update",
        target_type="national_push_config",
        target_id=str(config_row.id),
        reason=None,
        before_state=before_state,
        after_state={"enabled": config_row.enabled, "api_key_masked": _mask_api_key(config_row.api_key)},
    )
    db.commit()
    return NationalPushApiConfigOut(
        enabled=config_row.enabled,
        api_key_masked=_mask_api_key(config_row.api_key),
        updated_at=config_row.updated_at,
    )


def _execute_push_by_code(
    db: Session,
    record: NationalPushRecord,
    simulate_http_status: int | None,
) -> None:
    record.attempt_count = 0

    # 400: parameter validation issue
    if simulate_http_status == 400:
        record.attempt_count = 1
        record.failed_count = record.report_count
        record.pushed_count = 0
        record.status = "失败"
        record.last_error_code = 400
        record.last_error_message = "参数错误：请求参数校验失败"
        _append_push_log(db, record, 1, "失败", 400, "参数错误", {"hint": "请检查 period_code 或 payload 字段"})
        return

    # 401: API key authentication issue
    if simulate_http_status == 401:
        record.attempt_count = 1
        record.failed_count = record.report_count
        record.pushed_count = 0
        record.status = "失败"
        record.last_error_code = 401
        record.last_error_message = "认证失败：请检查 API Key"
        _append_push_log(db, record, 1, "失败", 401, "认证失败", {"hint": "API Key 无效或已过期"})
        return

    # 503: service unavailable, store for later retry
    if simulate_http_status == 503:
        record.attempt_count = 1
        record.failed_count = record.report_count
        record.pushed_count = 0
        record.status = "待重试"
        record.last_error_code = 503
        record.last_error_message = "目标系统不可用，已暂存待重试"
        _append_push_log(db, record, 1, "待重试", 503, "服务不可用，已暂存")
        return

    # 500: retry up to 3 times, then fail
    if simulate_http_status == 500:
        for attempt in range(1, NATIONAL_PUSH_MAX_RETRIES + 1):
            record.attempt_count = attempt
            if attempt < NATIONAL_PUSH_MAX_RETRIES:
                _append_push_log(db, record, attempt, "重试中", 500, f"第 {attempt} 次调用失败，准备重试")
            else:
                record.failed_count = record.report_count
                record.pushed_count = 0
                record.status = "失败"
                record.last_error_code = 500
                record.last_error_message = f"服务异常，已重试 {NATIONAL_PUSH_MAX_RETRIES} 次仍失败"
                _append_push_log(db, record, attempt, "失败", 500, record.last_error_message)
        return

    # default success
    record.attempt_count = 1
    record.failed_count = 0
    record.pushed_count = record.report_count
    record.status = "成功"
    record.last_error_code = None
    record.last_error_message = None
    _append_push_log(db, record, 1, "成功", 200, "推送成功")


@app.post("/national/push", response_model=NationalPushResultOut)
def national_push(
    payload: NationalPushExecuteIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> NationalPushResultOut:
    preview = _build_national_push_preview(db, payload.period_code)
    config_row = db.scalar(select(NationalPushApiConfig).order_by(NationalPushApiConfig.updated_at.desc()))

    batch_id = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{uuid4().hex[:6]}"
    record = NationalPushRecord(
        batch_id=batch_id,
        period_code=preview.period_code,
        report_count=preview.report_count,
        pushed_count=0,
        failed_count=0,
        status="重试中",
        attempt_count=0,
        request_payload={"period_code": preview.period_code, "payload": preview.payload},
        response_payload=None,
        created_by_user_id=current_user.id,
    )
    db.add(record)
    db.flush()

    if config_row is None or not config_row.enabled or not config_row.api_key.strip():
        record.attempt_count = 1
        record.status = "失败"
        record.failed_count = record.report_count
        record.last_error_code = 401
        record.last_error_message = "认证失败：API Key 未配置或已禁用"
        _append_push_log(db, record, 1, "失败", 401, "认证失败：API Key 未配置或已禁用")
        _write_audit_log(
            db=db,
            request=request,
            operator=current_user,
            action_type="national_push_execute",
            target_type="national_batch",
            target_id=batch_id,
            reason="API Key 未配置",
            before_state=None,
            after_state={"status": record.status, "failed_count": record.failed_count},
        )
        db.commit()
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="认证失败：请先配置可用 API Key")

    _execute_push_by_code(db, record, payload.simulate_http_status)
    record.response_payload = {
        "status": record.status,
        "attempt_count": record.attempt_count,
        "failed_count": record.failed_count,
        "last_error_code": record.last_error_code,
    }

    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="national_push_execute",
        target_type="national_batch",
        target_id=batch_id,
        reason=None,
        before_state=None,
        after_state=record.response_payload,
    )
    db.commit()

    if payload.simulate_http_status == 400:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="参数错误：推送参数非法")
    if payload.simulate_http_status == 401:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="认证失败：请检查 API Key")

    return NationalPushResultOut(
        batch_id=record.batch_id,
        pushed_count=record.pushed_count,
        failed_count=record.failed_count,
        status=record.status,
        attempt_count=record.attempt_count,
    )


@app.get("/national/push/records", response_model=list[NationalPushRecordOut])
def national_push_records(
    status_value: str | None = None,
    batch_id: str | None = None,
    period_code: str | None = None,
    limit: int = 100,
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> list[NationalPushRecordOut]:
    stmt = select(NationalPushRecord).order_by(NationalPushRecord.created_at.desc())
    if status_value:
        stmt = stmt.where(NationalPushRecord.status == status_value)
    if batch_id:
        stmt = stmt.where(NationalPushRecord.batch_id.contains(batch_id.strip()))
    if period_code:
        stmt = stmt.where(NationalPushRecord.period_code == period_code)
    safe_limit = min(max(limit, 1), 500)
    stmt = stmt.limit(safe_limit)
    return list(db.scalars(stmt).all())


@app.get("/national/push/records/{batch_id}/logs", response_model=list[NationalPushLogOut])
def national_push_record_logs(
    batch_id: str,
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> list[NationalPushLogOut]:
    record = db.scalar(select(NationalPushRecord).where(NationalPushRecord.batch_id == batch_id))
    if record is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="推送批次不存在")
    stmt = select(NationalPushLog).where(NationalPushLog.batch_id == batch_id).order_by(NationalPushLog.created_at.asc())
    return list(db.scalars(stmt).all())


@app.post("/national/push/records/{batch_id}/retry", response_model=NationalPushResultOut)
def national_push_retry(
    batch_id: str,
    payload: NationalPushExecuteIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> NationalPushResultOut:
    record = db.scalar(select(NationalPushRecord).where(NationalPushRecord.batch_id == batch_id))
    if record is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="推送批次不存在")
    if record.status not in {"待重试", "失败"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="仅失败或待重试批次可重试")

    _execute_push_by_code(db, record, payload.simulate_http_status)
    record.updated_at = datetime.utcnow()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="national_push_retry",
        target_type="national_batch",
        target_id=batch_id,
        reason=None,
        before_state=None,
        after_state={"status": record.status, "attempt_count": record.attempt_count},
    )
    db.commit()
    if payload.simulate_http_status == 400:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="参数错误：推送参数非法")
    if payload.simulate_http_status == 401:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="认证失败：请检查 API Key")
    return NationalPushResultOut(
        batch_id=record.batch_id,
        pushed_count=record.pushed_count,
        failed_count=record.failed_count,
        status=record.status,
        attempt_count=record.attempt_count,
    )


@app.get("/notices", response_model=list[NoticeOut])
def list_notices(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[NoticeOut]:
    stmt = select(Notice).where(_notice_visible_filter(current_user)).order_by(Notice.created_at.desc())
    notices = list(db.scalars(stmt).all())
    if not notices:
        return []

    notice_ids = [item.id for item in notices]
    read_ids = set(
        db.scalars(
            select(NoticeRead.notice_id).where(
                NoticeRead.user_id == current_user.id,
                NoticeRead.notice_id.in_(notice_ids),
            )
        ).all()
    )
    return [_notice_to_out(item, item.id in read_ids) for item in notices]


@app.get("/notices/unread-count")
def notice_unread_count(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict[str, int]:
    visible_notice_ids = list(
        db.scalars(select(Notice.id).where(_notice_visible_filter(current_user))).all()
    )
    if not visible_notice_ids:
        return {"unread_count": 0}

    read_count = db.scalar(
        select(func.count(NoticeRead.id)).where(
            NoticeRead.user_id == current_user.id,
            NoticeRead.notice_id.in_(visible_notice_ids),
        )
    ) or 0
    return {"unread_count": max(len(visible_notice_ids) - read_count, 0)}


@app.get("/notices/{notice_id}", response_model=NoticeOut)
def notice_detail(
    notice_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> NoticeOut:
    notice = db.get(Notice, notice_id)
    if notice is None or notice.is_deleted:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="通知不存在")

    visible = db.scalar(select(Notice.id).where(Notice.id == notice_id, _notice_visible_filter(current_user)))
    if visible is None:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权限查看该通知")

    is_read = db.scalar(
        select(func.count(NoticeRead.id)).where(
            NoticeRead.notice_id == notice_id,
            NoticeRead.user_id == current_user.id,
        )
    )
    return _notice_to_out(notice, bool(is_read))


@app.post("/notices/{notice_id}/read")
def mark_notice_read(
    notice_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    notice = db.get(Notice, notice_id)
    if notice is None or notice.is_deleted:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="通知不存在")

    visible = db.scalar(select(Notice.id).where(Notice.id == notice_id, _notice_visible_filter(current_user)))
    if visible is None:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权限查看该通知")

    existing = db.scalar(
        select(NoticeRead).where(
            NoticeRead.notice_id == notice_id,
            NoticeRead.user_id == current_user.id,
        )
    )
    if existing is None:
        db.add(NoticeRead(notice_id=notice_id, user_id=current_user.id))
        notice.read_count += 1
        _write_audit_log(
            db=db,
            request=request,
            operator=current_user,
            action_type="notice_read",
            target_type="notice",
            target_id=str(notice_id),
            reason=None,
            before_state=None,
            after_state={"read": True},
        )
        db.commit()

    return {"read": True}


@app.get("/province/notices/manage", response_model=list[NoticeOut])
def province_manage_notices(
    _: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> list[NoticeOut]:
    notices = list(
        db.scalars(
            select(Notice)
            .where(Notice.scope == NoticeScope.PROVINCE, Notice.is_deleted.is_(False))
            .order_by(Notice.created_at.desc())
        ).all()
    )
    return [_notice_to_out(item, False) for item in notices]


@app.post("/province/notices", response_model=NoticeOut)
def create_province_notice(
    payload: NoticeCreateIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> NoticeOut:
    notice = Notice(
        title=payload.title,
        content=payload.content,
        scope=NoticeScope.PROVINCE,
        city_code=None,
        publisher_user_id=current_user.id,
        publisher_name=current_user.username,
        publisher_role=current_user.role.value,
    )
    db.add(notice)
    db.flush()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="notice_create",
        target_type="notice",
        target_id=str(notice.id),
        reason=None,
        before_state=None,
        after_state={"title": notice.title, "scope": notice.scope.value},
    )
    db.commit()
    db.refresh(notice)
    return _notice_to_out(notice, False)


@app.put("/province/notices/{notice_id}", response_model=NoticeOut)
def update_province_notice(
    notice_id: int,
    payload: NoticeUpdateIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> NoticeOut:
    notice = db.get(Notice, notice_id)
    if notice is None or notice.is_deleted or notice.scope != NoticeScope.PROVINCE:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="通知不存在")

    before_state = {"title": notice.title, "content": notice.content}
    notice.title = payload.title
    notice.content = payload.content

    db.flush()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="notice_update",
        target_type="notice",
        target_id=str(notice.id),
        reason=None,
        before_state=before_state,
        after_state={"title": notice.title, "content": notice.content},
    )
    db.commit()
    db.refresh(notice)
    return _notice_to_out(notice, False)


@app.delete("/province/notices/{notice_id}")
def delete_province_notice(
    notice_id: int,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.PROVINCE)),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    notice = db.get(Notice, notice_id)
    if notice is None or notice.is_deleted or notice.scope != NoticeScope.PROVINCE:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="通知不存在")

    before_state = {"title": notice.title, "content": notice.content, "is_deleted": notice.is_deleted}
    notice.is_deleted = True
    notice.deleted_at = datetime.utcnow()

    db.flush()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="notice_delete",
        target_type="notice",
        target_id=str(notice.id),
        reason=None,
        before_state=before_state,
        after_state={"is_deleted": True},
    )
    db.commit()
    return {"deleted": True}


@app.get("/city/notices/manage", response_model=list[NoticeOut])
def city_manage_notices(
    current_user: User = Depends(require_roles(UserRole.CITY)),
    db: Session = Depends(get_db),
) -> list[NoticeOut]:
    notices = list(
        db.scalars(
            select(Notice)
            .where(
                Notice.scope == NoticeScope.CITY,
                Notice.city_code == current_user.city_code,
                Notice.publisher_user_id == current_user.id,
                Notice.is_deleted.is_(False),
            )
            .order_by(Notice.created_at.desc())
        ).all()
    )
    return [_notice_to_out(item, False) for item in notices]


@app.post("/city/notices", response_model=NoticeOut)
def create_city_notice(
    payload: NoticeCreateIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.CITY)),
    db: Session = Depends(get_db),
) -> NoticeOut:
    if not current_user.city_code:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="当前市级账号未绑定辖区")

    notice = Notice(
        title=payload.title,
        content=payload.content,
        scope=NoticeScope.CITY,
        city_code=current_user.city_code,
        publisher_user_id=current_user.id,
        publisher_name=current_user.username,
        publisher_role=current_user.role.value,
    )
    db.add(notice)
    db.flush()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="notice_create",
        target_type="notice",
        target_id=str(notice.id),
        reason=None,
        before_state=None,
        after_state={"title": notice.title, "scope": notice.scope.value, "city_code": notice.city_code},
    )
    db.commit()
    db.refresh(notice)
    return _notice_to_out(notice, False)


@app.put("/city/notices/{notice_id}", response_model=NoticeOut)
def update_city_notice(
    notice_id: int,
    payload: NoticeUpdateIn,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.CITY)),
    db: Session = Depends(get_db),
) -> NoticeOut:
    notice = db.get(Notice, notice_id)
    if (
        notice is None
        or notice.is_deleted
        or notice.scope != NoticeScope.CITY
        or notice.city_code != current_user.city_code
        or notice.publisher_user_id != current_user.id
    ):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="通知不存在")

    before_state = {"title": notice.title, "content": notice.content}
    notice.title = payload.title
    notice.content = payload.content
    db.flush()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="notice_update",
        target_type="notice",
        target_id=str(notice.id),
        reason=None,
        before_state=before_state,
        after_state={"title": notice.title, "content": notice.content},
    )
    db.commit()
    db.refresh(notice)
    return _notice_to_out(notice, False)


@app.delete("/city/notices/{notice_id}")
def delete_city_notice(
    notice_id: int,
    request: Request,
    current_user: User = Depends(require_roles(UserRole.CITY)),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    notice = db.get(Notice, notice_id)
    if (
        notice is None
        or notice.is_deleted
        or notice.scope != NoticeScope.CITY
        or notice.city_code != current_user.city_code
        or notice.publisher_user_id != current_user.id
    ):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="通知不存在")

    before_state = {"title": notice.title, "content": notice.content, "is_deleted": notice.is_deleted}
    notice.is_deleted = True
    notice.deleted_at = datetime.utcnow()
    db.flush()
    _write_audit_log(
        db=db,
        request=request,
        operator=current_user,
        action_type="notice_delete",
        target_type="notice",
        target_id=str(notice.id),
        reason=None,
        before_state=before_state,
        after_state={"is_deleted": True},
    )
    db.commit()
    return {"deleted": True}
